#!/usr/bin/env python3
from __future__ import annotations

import html
import json
import calendar
import mimetypes
import socket
import sqlite3
import os
import sys
import subprocess
import threading
import webbrowser
import datetime as dt
import uuid
import zipfile
import re
from io import BytesIO
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, quote, urlencode, urlparse

from openpyxl import load_workbook

from fit_to_excel import (
    APP_VERSION,
    DEFAULT_OUTPUT_DIR,
    DROPDOWN_CONFIG_PATH,
    WORKBOOK_VERSION_NAME,
    create_workbook,
    default_output_path,
    load_dropdown_options,
    output_month_label,
    weighted_average,
    write_fit_to_sqlite,
)


ROOT = Path(__file__).resolve().parent
FIT_DIR = ROOT / "FIT"
ASSETS_DIR = ROOT / "assets"
GARMIN_CONFIG_PATH = ROOT / "config" / "garmin_connect.json"
SQLITE_DB_PATH = ROOT / "analysis_platform" / "running_analytics.sqlite"
HOST = "127.0.0.1"
PORT = 8765
PLATFORM_URL = "http://127.0.0.1:8766/"
EXCEL_FORMAT_VERSION = WORKBOOK_VERSION_NAME
EXCEL_SCHEMA_LABEL = EXCEL_FORMAT_VERSION.replace("跑步分析資料 ", "Excel Schema ")
DEFAULT_FIT_LIST_LIMIT = 30
DOWNLOAD_JOBS = {}
DOWNLOAD_JOBS_LOCK = threading.Lock()
BATCH_JOBS = {}
BATCH_JOBS_LOCK = threading.Lock()
OPTION_FIELDS = [
    ("shoes", "鞋款"),
    ("workout_types", "課表類型"),
    ("training_focus", "訓練目的"),
    ("garmin_rpe", "感受難度"),
    ("garmin_feel", "感覺如何"),
]
ACTIVITY_INFO_FIELDS = [
    ("activity_name", "活動名稱", "text", "活動名稱"),
    ("shoe", "鞋款", "select", "鞋款"),
    ("workout_type", "課表類型", "select", "課表類型"),
    ("training_focus", "訓練目的", "multi", "訓練目的"),
    ("weather_temp", "天氣氣溫 (°C)", "number", "氣溫(°C)"),
    ("humidity", "濕度 (%)", "number", "濕度(%)"),
    ("wind_direction", "風向", "text", "風向"),
    ("wind_speed", "風速", "text", "風速"),
    ("weather_description", "天氣描述", "text", "天氣描述"),
    ("feel", "感覺如何", "select", "感覺如何"),
    ("rpe", "感受難度", "select", "感受難度"),
    ("fueling", "補給紀錄", "textarea", "補給紀錄"),
    ("notes", "備註", "textarea", "備註"),
    ("max_hr", "最大心率", "number", "最大心率"),
    ("critical_power", "Critical Power (W)", "number", "Critical Power(W)"),
    ("training_effect_aerobic", "Training Effect (Aerobic)", "number", "Training Effect (Aerobic)"),
    ("training_effect_anaerobic", "Training Effect (Anaerobic)", "number", "Training Effect (Anaerobic)"),
    ("training_load", "Training Load", "number", "Training Load"),
    ("recovery_time_hr", "Recovery Time (hr)", "number", "Recovery Time (hr)"),
]
WORKOUT_FOCUS_MAP_KEY = "workout_focus_map"
DEFAULT_WORKOUT_FOCUS_HINTS = {
    "Recovery": ["Recovery"],
    "Easy": ["Aerobic Base"],
    "LSD": ["Endurance"],
    "Long Run": ["Endurance"],
    "Progression": ["Endurance"],
    "Tempo": ["Threshold"],
    "Marathon Pace": ["Marathon Pace"],
    "Interval": ["VO₂max"],
    "Repetition": ["Speed"],
    "Hill": ["Running Economy"],
    "Race": ["Race Simulation", "Test"],
    "Test": ["Race Simulation", "Test"],
    "Fartlek": ["Speed", "VO₂max"],
}


def open_file(path):
    if sys.platform.startswith("win"):
        os.startfile(path)  # type: ignore[attr-defined]
    elif sys.platform == "darwin":
        subprocess.run(["open", str(path)], check=False)
    else:
        subprocess.run(["xdg-open", str(path)], check=False)


def is_output_file(path):
    try:
        resolved = path.resolve()
        output_root = DEFAULT_OUTPUT_DIR.resolve()
        return resolved.is_file() and resolved.suffix.lower() == ".xlsx" and resolved.is_relative_to(output_root)
    except OSError:
        return False


def is_output_dir(path):
    try:
        resolved = path.resolve()
        output_root = DEFAULT_OUTPUT_DIR.resolve()
        return resolved.is_dir() and resolved.is_relative_to(output_root)
    except OSError:
        return False


def pace_text(seconds, meters):
    if not seconds or not meters:
        return ""
    sec_per_km = round(float(seconds) / float(meters) * 1000)
    return f"{sec_per_km // 60}:{sec_per_km % 60:02d}/km"


def format_duration(seconds):
    if not seconds:
        return ""
    seconds = int(round(float(seconds)))
    hours, remainder = divmod(seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    if hours:
        return f"{hours}:{minutes:02d}:{seconds:02d}"
    return f"{minutes}:{seconds:02d}"


def cell_map(ws):
    return {ws.cell(row, 1).value: ws.cell(row, 2).value for row in range(1, ws.max_row + 1)}


def load_raw_config():
    if not DROPDOWN_CONFIG_PATH.exists():
        return {}
    try:
        return json.loads(DROPDOWN_CONFIG_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}


def load_garmin_config():
    if not GARMIN_CONFIG_PATH.exists():
        return {"email": ""}
    try:
        loaded = json.loads(GARMIN_CONFIG_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {"email": ""}
    return {
        "email": str(loaded.get("email", "")),
        "password": str(loaded.get("password", "")),
    }


def save_garmin_config(email, password):
    GARMIN_CONFIG_PATH.parent.mkdir(parents=True, exist_ok=True)
    GARMIN_CONFIG_PATH.write_text(
        json.dumps({"email": email, "password": password}, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def remove_garmin_config():
    try:
        GARMIN_CONFIG_PATH.unlink()
    except FileNotFoundError:
        pass


def matching_option(options, hint):
    hint = hint.lower()
    for option in options:
        if hint in str(option).lower():
            return option
    return None


def default_workout_focus_map(options):
    result = {}
    training_focus = options.get("training_focus", [])
    for workout in options.get("workout_types", []):
        matched_focus = []
        for workout_hint, focus_hints in DEFAULT_WORKOUT_FOCUS_HINTS.items():
            if workout_hint.lower() not in str(workout).lower():
                continue
            for focus_hint in focus_hints:
                focus = matching_option(training_focus, focus_hint)
                if focus and focus not in matched_focus:
                    matched_focus.append(focus)
            break
        result[workout] = matched_focus
    return result


def clean_workout_focus_map(raw_map, options):
    workout_types = options.get("workout_types", [])
    training_focus = set(options.get("training_focus", []))
    result = {}
    if isinstance(raw_map, dict):
        for workout in workout_types:
            values = raw_map.get(workout, [])
            if isinstance(values, str):
                values = [values]
            if isinstance(values, list):
                result[workout] = [
                    str(value).strip()
                    for value in values
                    if str(value).strip() in training_focus
                ]
    defaults = default_workout_focus_map(options)
    for workout in workout_types:
        result.setdefault(workout, defaults.get(workout, []))
    return result


def shoe_label_from_row(row):
    model = str(row["model"] or "").strip()
    nickname = str(row["nickname"] or "").strip()
    if model and nickname:
        return f"{model} {nickname}".strip()
    if model:
        return model
    if nickname:
        return nickname
    return str(row["shoe_code"] or "").strip()


def shoe_labels_from_sqlite():
    if not SQLITE_DB_PATH.exists():
        return []
    try:
        connection = sqlite3.connect(SQLITE_DB_PATH)
        connection.row_factory = sqlite3.Row
        rows = connection.execute(
            """
            SELECT shoe_code, model, nickname, is_active
            FROM shoe
            ORDER BY is_active DESC, model, nickname, shoe_code
            """
        ).fetchall()
    except sqlite3.DatabaseError:
        return []
    finally:
        try:
            connection.close()
        except Exception:
            pass

    labels = []
    seen = set()
    for row in rows:
        label = shoe_label_from_row(row)
        key = label.lower()
        if not label or key in seen:
            continue
        seen.add(key)
        labels.append(label)
    return labels


def load_app_options():
    options = load_dropdown_options(DROPDOWN_CONFIG_PATH)
    raw = load_raw_config()
    if not raw.get("shoes"):
        sqlite_shoes = shoe_labels_from_sqlite()
        if sqlite_shoes:
            options["shoes"] = sqlite_shoes
            updated = dict(raw)
            updated["shoes"] = sqlite_shoes
            save_dropdown_options(updated)
    options[WORKOUT_FOCUS_MAP_KEY] = clean_workout_focus_map(raw.get(WORKOUT_FOCUS_MAP_KEY), options)
    return options


def workbook_summary(path):
    wb = load_workbook(path, data_only=True)
    info = cell_map(wb["活動資訊"])
    km = wb["每公里數據"]
    rows = [
        row
        for row in km.iter_rows(min_row=3, max_row=km.max_row, values_only=True)
        if row and isinstance(row[0], int)
    ]
    total_distance = sum(row[1] or 0 for row in rows)
    total_seconds = sum(row[2] or 0 for row in rows)
    avg_hr_pairs = [
        (row[4], row[2])
        for row in rows
        if isinstance(row[4], (int, float)) and isinstance(row[2], (int, float))
    ]
    avg_power_pairs = [
        (row[8], row[2])
        for row in rows
        if isinstance(row[8], (int, float)) and isinstance(row[2], (int, float))
    ]

    summary = [
        ("活動日期", info.get("活動日期")),
        ("開始時間", info.get("開始時間")),
        ("距離", f"{total_distance / 1000:.2f} km" if total_distance else ""),
        ("時間", format_duration(total_seconds)),
        ("平均配速", pace_text(total_seconds, total_distance)),
        ("平均心率", weighted_average(avg_hr_pairs, 1) if avg_hr_pairs else ""),
        ("平均功率", f"{weighted_average(avg_power_pairs, 1)} W" if avg_power_pairs else ""),
        ("天氣", weather_summary(info)),
        ("Training Effect", training_effect_summary(info)),
        ("Training Load", info.get("Training Load")),
    ]
    return [(label, value) for label, value in summary if value not in ("", None)]


def activity_info_row_map(ws):
    return {
        str(ws.cell(row, 1).value): row
        for row in range(1, ws.max_row + 1)
        if ws.cell(row, 1).value not in ("", None)
    }


def load_activity_info_values(path):
    wb = load_workbook(path, data_only=False)
    if "活動資訊" not in wb.sheetnames:
        raise ValueError("這個 Excel 沒有「活動資訊」工作表。")
    ws = wb["活動資訊"]
    rows = activity_info_row_map(ws)
    values = {}
    for key, label, _field_type, _display in ACTIVITY_INFO_FIELDS:
        row = rows.get(label)
        values[key] = ws.cell(row, 2).value if row else ""
    return {key: "" if value is None else value for key, value in values.items()}


def save_activity_info_values(path, values):
    wb = load_workbook(path)
    if "活動資訊" not in wb.sheetnames:
        raise ValueError("這個 Excel 沒有「活動資訊」工作表。")
    ws = wb["活動資訊"]
    rows = activity_info_row_map(ws)
    for key, label, _field_type, _display in ACTIVITY_INFO_FIELDS:
        row = rows.get(label)
        if not row:
            continue
        ws.cell(row, 2, values.get(key, ""))
    wb.save(path)


def activity_info_values_from_form(form):
    values = {}
    for key, _label, field_type, _display in ACTIVITY_INFO_FIELDS:
        if field_type == "multi":
            values[key] = "、".join(selected_values(form, key))
        else:
            values[key] = parse_number(first_value(form, key))
    return values


def weather_summary(info):
    temp = info.get("天氣氣溫 (°C)", info.get("天氣氣溫(°C)"))
    humidity = info.get("濕度 (%)", info.get("濕度(%)"))
    wind_direction = info.get("風向")
    wind_speed = info.get("風速")
    weather_description = info.get("天氣描述")
    parts = []
    if weather_description not in ("", None):
        parts.append(str(weather_description))
    if temp not in ("", None):
        parts.append(f"{temp}°C")
    if humidity not in ("", None):
        parts.append(f"{humidity}%")
    wind = " ".join(str(value) for value in (wind_direction, wind_speed) if value not in ("", None))
    if wind:
        parts.append(wind)
    return " / ".join(parts)


def training_effect_summary(info):
    aerobic = info.get("Training Effect (Aerobic)")
    anaerobic = info.get("Training Effect (Anaerobic)")
    parts = []
    if aerobic not in ("", None):
        parts.append(f"Aerobic {aerobic}")
    if anaerobic not in ("", None):
        parts.append(f"Anaerobic {anaerobic}")
    return " / ".join(parts)


def summary_html(items):
    if not items:
        return ""
    rows = "\n".join(
        f"<tr><th>{html.escape(str(label))}</th><td>{html.escape(str(value))}</td></tr>"
        for label, value in items
    )
    return f"""
      <table class="summary">
        <tbody>
          {rows}
        </tbody>
      </table>
    """


def friendly_error(error):
    text = str(error)
    if isinstance(error, FileNotFoundError) and "SQLite Schema v1.0.sql" in text:
        return "找不到 SQLite Schema 檔案，請確認專案文件路徑完整，或重新啟動應用程式後再試一次。"
    if isinstance(error, FileNotFoundError):
        return "找不到指定的檔案，請確認 FIT 檔還在原本的位置。"
    if isinstance(error, PermissionError):
        return "目前沒有權限讀寫這個檔案或資料夾，請確認檔案沒有被 Excel 開著，或換一個輸出檔名再試一次。"
    if isinstance(error, ModuleNotFoundError):
        return "缺少必要套件，請重新啟動應用程式，讓啟動檔自動安裝需求套件。"
    if isinstance(error, socket.timeout):
        return "天氣查詢逾時。可以稍後再試，或先取消自動抓天氣完成轉檔。"
    if "No lap data found" in text:
        return "這個 FIT 裡沒有可用的每公里分段資料，可能不是跑步活動，或檔案內容不完整。"
    if "FIT decode errors" in text:
        return "FIT 檔解析失敗，請確認這是 Garmin Connect 匯出的 Original FIT 檔。"
    if "urlopen" in text or "Open-Meteo" in text:
        return "天氣查詢失敗。請確認網路可用，或先取消自動抓天氣完成轉檔。"
    return f"轉檔失敗：{text}"


def friendly_download_error(error):
    text = str(error)
    if isinstance(error, ModuleNotFoundError):
        return "缺少 Garmin Connect 下載套件，請重新啟動應用程式，讓啟動檔自動安裝需求套件。"
    if "401" in text or "Unauthorized" in text or "Bad credentials" in text:
        return "Garmin 登入失敗，請確認帳號密碼是否正確。"
    if "MFA" in text or "two" in text.lower() or "verification" in text.lower():
        return "Garmin 需要額外驗證。請先用瀏覽器登入 Garmin Connect 完成驗證後再試一次。"
    if "429" in text:
        return "Garmin 暫時限制請求次數，請稍後再試。"
    return f"下載失敗：{text}"


def all_fit_files():
    FIT_DIR.mkdir(parents=True, exist_ok=True)
    return sorted(FIT_DIR.rglob("*.fit"), key=lambda path: path.stat().st_mtime, reverse=True)


def relative_fit_path(path):
    try:
        return path.resolve().relative_to(FIT_DIR.resolve()).as_posix()
    except ValueError:
        return path.name


def resolve_fit_path(value):
    if not value:
        return FIT_DIR / ""
    candidate = (FIT_DIR / value).resolve()
    fit_root = FIT_DIR.resolve()
    try:
        candidate.relative_to(fit_root)
    except ValueError:
        return FIT_DIR / ""
    return candidate


def fit_month_label(path):
    relative = relative_fit_path(path)
    first_part = relative.split("/", 1)[0]
    if re.fullmatch(r"\d{4}-\d{2}", first_part):
        return first_part

    try:
        return output_month_label(path)
    except Exception:
        pass
    return dt.date.today().strftime("%Y-%m")


def excel_output_dir_for_fit(path):
    return excel_output_path_for_fit(path).parent


def excel_output_path_for_fit(path):
    return default_output_path(path)


def is_running_activity(activity):
    activity_type = activity.get("activityType") or {}
    type_key = str(activity_type.get("typeKey") or "").lower()
    return "running" in type_key


def activity_file_name(activity):
    activity_id = activity.get("activityId")
    start_time = str(activity.get("startTimeLocal") or activity.get("startTimeGMT") or "")
    date_part = start_time.split()[0].replace("-", "") if start_time else "garmin"
    if not date_part:
        date_part = "garmin"
    return f"{date_part}_{activity_id}_ACTIVITY.fit"


def normalize_fit_download(data):
    if isinstance(data, bytes):
        raw = data
    if isinstance(data, bytearray):
        raw = bytes(data)
    if isinstance(data, str):
        raw = data.encode("utf-8")
    if not isinstance(data, (bytes, bytearray, str)):
        raise ValueError("Garmin 回傳的 FIT 檔案格式無法寫入。")

    if raw.startswith(b"PK"):
        with zipfile.ZipFile(BytesIO(raw)) as archive:
            fit_names = [name for name in archive.namelist() if name.lower().endswith(".fit")]
            if not fit_names:
                raise ValueError("Garmin 回傳的壓縮檔裡沒有 FIT 檔。")
            return archive.read(fit_names[0])
    return raw


def download_running_fit_by_date_range(email, password, start_date, end_date, target_month, progress=None):
    from garminconnect import Garmin

    target_dir = FIT_DIR / target_month

    if progress:
        progress("登入 Garmin Connect...", 0, 0, "running")

    client = Garmin(email, password)
    client.login()

    if progress:
        progress(f"搜尋 {start_date} 到 {end_date} 的活動...", 0, 0, "running")

    activities = client.get_activities_by_date(start_date, end_date)
    running_activities = [activity for activity in activities if is_running_activity(activity)]

    target_dir.mkdir(parents=True, exist_ok=True)
    downloaded = []
    skipped = []
    for activity in running_activities:
        current = len(downloaded) + len(skipped) + 1
        target = target_dir / activity_file_name(activity)
        if progress:
            progress(f"處理中：{target.name}", current - 1, len(running_activities), "running")
        if target.exists():
            skipped.append(target)
            if progress:
                progress(f"已存在，略過：{target.name}", current, len(running_activities), "running")
            continue
        data = normalize_fit_download(
            client.download_activity(
                activity.get("activityId"),
                dl_fmt=Garmin.ActivityDownloadFormat.ORIGINAL,
            )
        )
        target.write_bytes(data)
        downloaded.append(target)
        if progress:
            progress(f"已下載：{target.name}", current, len(running_activities), "running")

    if progress:
        progress("整理下載結果...", len(running_activities), len(running_activities), "running")

    return {
        "start_date": start_date,
        "end_date": end_date,
        "total": len(running_activities),
        "downloaded": downloaded,
        "skipped": skipped,
        "target_dir": target_dir,
    }


def download_monthly_running_fit(email, password, year, month, progress=None):
    _, last_day = calendar.monthrange(year, month)
    start_date = f"{year:04d}-{month:02d}-01"
    end_date = f"{year:04d}-{month:02d}-{last_day:02d}"
    return download_running_fit_by_date_range(
        email,
        password,
        start_date,
        end_date,
        f"{year:04d}-{month:02d}",
        progress=progress,
    )


def download_today_running_fit(email, password, progress=None):
    today = dt.date.today()
    date_text = today.isoformat()
    return download_running_fit_by_date_range(
        email,
        password,
        date_text,
        date_text,
        today.strftime("%Y-%m"),
        progress=progress,
    )


def fit_files(selected_fit=""):
    files = all_fit_files()
    total_count = len(files)
    limit = DEFAULT_FIT_LIST_LIMIT
    limited = files[:limit]
    if selected_fit and not any(relative_fit_path(path) == selected_fit for path in limited):
        selected_path = resolve_fit_path(selected_fit)
        if selected_path.exists() and selected_path.suffix.lower() == ".fit":
            limited.insert(0, selected_path)
    return limited, total_count, limit


def available_fit_months():
    months = {fit_month_label(path) for path in all_fit_files()}
    return sorted(months, reverse=True)


def fit_files_for_month(month):
    return sorted(
        [path for path in all_fit_files() if fit_month_label(path) == month],
        key=lambda path: relative_fit_path(path),
    )


def batch_file_status(fit_path):
    output_path = excel_output_path_for_fit(fit_path)
    if not output_path.exists():
        return "未轉檔", output_path
    try:
        if fit_path.stat().st_mtime > output_path.stat().st_mtime:
            return "FIT 較新", output_path
    except OSError:
        pass
    return "已轉檔", output_path


def batch_scan(month):
    rows = []
    for fit_path in fit_files_for_month(month):
        status, output_path = batch_file_status(fit_path)
        rows.append(
            {
                "fit_path": fit_path,
                "fit": relative_fit_path(fit_path),
                "output_path": output_path,
                "output": output_path.name,
                "status": status,
            }
        )
    return rows


def batch_status_table(month):
    rows = batch_scan(month)
    if not rows:
        return '<p class="note">這個月份目前沒有 FIT 檔。</p>'
    html_rows = []
    for row in rows:
        html_rows.append(
            f"""
            <tr>
              <td><code>{html.escape(row["fit"])}</code></td>
              <td>{html.escape(row["status"])}</td>
              <td><code>{html.escape(row["output"])}</code></td>
            </tr>
            """
        )
    return f"""
      <div class="reference-table-wrap">
        <table class="reference-table">
          <thead>
            <tr>
              <th>FIT</th>
              <th>狀態</th>
              <th>Excel</th>
            </tr>
          </thead>
          <tbody>
            {"".join(html_rows)}
          </tbody>
        </table>
      </div>
    """


def update_download_job(job_id, **updates):
    with DOWNLOAD_JOBS_LOCK:
        job = DOWNLOAD_JOBS.get(job_id)
        if not job:
            return
        job.update(updates)


def download_job_result_html(job):
    result = job.get("result") or {}
    downloaded = result.get("downloaded") or []
    skipped = result.get("skipped") or []
    downloaded_count = len(downloaded)
    skipped_count = len(skipped)
    target_dir = result.get("target_dir", "")
    folder_fit_count = 0
    if target_dir:
        try:
            folder_fit_count = len(list(Path(target_dir).glob("*.fit")))
        except OSError:
            folder_fit_count = 0
    latest_names = [relative_fit_path(path) for path in downloaded[:8]]
    if latest_names:
        downloaded_list = "<br>".join(f"<code>{html.escape(name)}</code>" for name in latest_names)
        if downloaded_count > len(latest_names):
            downloaded_list += f"<br>...另有 {downloaded_count - len(latest_names)} 個檔案"
    else:
        downloaded_list = "沒有新增檔案"
    target_month = job.get("target_month", "")
    return (
        f"{html.escape(target_month)} 下載完成：找到 {result.get('total', 0)} 個跑步活動，"
        f"新增 {downloaded_count} 個，已存在略過 {skipped_count} 個。<br>"
        f"儲存位置：<code>{html.escape(str(target_dir))}</code><br>"
        f"此月份資料夾目前共有 {folder_fit_count} 個 FIT 檔。<br>"
        f"{downloaded_list}<br><br>"
        f'<a class="button" href="/">回轉檔頁</a> '
        f'<a class="button secondary" href="/download-fit">下載其他月份</a>'
    )


def download_job_snapshot(job_id):
    with DOWNLOAD_JOBS_LOCK:
        job = dict(DOWNLOAD_JOBS.get(job_id) or {})
    if not job:
        return {"found": False}
    current = int(job.get("current", 0) or 0)
    total = int(job.get("total", 0) or 0)
    percent = round(current / total * 100) if total else 0
    snapshot = {
        "found": True,
        "status": job.get("status", "running"),
        "message": job.get("message", ""),
        "current": current,
        "total": total,
        "percent": percent,
        "error": job.get("error", ""),
        "result_html": "",
    }
    if snapshot["status"] == "done" and "result" in job:
        snapshot["result_html"] = download_job_result_html(job)
    elif snapshot["status"] == "done":
        snapshot["status"] = "running"
        snapshot["message"] = "整理下載結果..."
    return snapshot


def run_download_job(job_id, email, password, mode, year, month):
    def progress(message, current, total, status):
        update_download_job(
            job_id,
            message=message,
            current=current,
            total=total,
            status=status,
        )

    try:
        if mode == "today":
            result = download_today_running_fit(email, password, progress=progress)
        else:
            result = download_monthly_running_fit(email, password, year, month, progress=progress)
    except Exception as error:
        update_download_job(
            job_id,
            status="error",
            error=friendly_download_error(error),
            message="下載中斷。",
        )
        return
    update_download_job(
        job_id,
        status="done",
        result=result,
        current=result["total"],
        total=result["total"],
        message="下載完成。",
    )


def start_download_job(email, password, year, month, mode="month"):
    target_month = dt.date.today().strftime("%Y-%m") if mode == "today" else f"{year:04d}-{month:02d}"
    job_id = uuid.uuid4().hex
    with DOWNLOAD_JOBS_LOCK:
        DOWNLOAD_JOBS[job_id] = {
            "status": "running",
            "message": "準備下載...",
            "current": 0,
            "total": 0,
            "target_month": target_month,
        }
    worker = threading.Thread(
        target=run_download_job,
        args=(job_id, email, password, mode, year, month),
        daemon=True,
    )
    worker.start()
    return job_id


def update_batch_job(job_id, **updates):
    with BATCH_JOBS_LOCK:
        job = BATCH_JOBS.get(job_id)
        if not job:
            return
        job.update(updates)


def should_batch_convert(row, overwrite_newer=False, overwrite_all=False):
    if overwrite_all:
        return True
    if row["status"] == "未轉檔":
        return True
    if overwrite_newer and row["status"] == "FIT 較新":
        return True
    return False


def batch_convert_result_html(job, job_id=""):
    result = job.get("result") or {}
    month = html.escape(job.get("month", ""))
    success = result.get("success", [])
    skipped = result.get("skipped", [])
    failed = result.get("failed", [])
    edited = set(job.get("edited", []))
    output_dir = DEFAULT_OUTPUT_DIR / job.get("month", "")
    parts = [
        f"{month} 批次轉檔完成：成功 {len(success)} 個，略過 {len(skipped)} 個，失敗 {len(failed)} 個。<br>",
        f"輸出位置：<code>{html.escape(str(output_dir))}</code><br>",
    ]
    if failed:
        parts.append("<br>失敗項目：<br>")
        for item in failed[:12]:
            parts.append(f"<code>{html.escape(item['fit'])}</code>：{html.escape(item['error'])}<br>")
        if len(failed) > 12:
            parts.append(f"...另有 {len(failed) - 12} 個失敗項目<br>")
    if success:
        parts.append("<br>這次轉好的 Excel：<br>")
        for saved in success[:20]:
            edit_params = {"path": saved}
            if job_id:
                edit_params["job"] = job_id
            edit_link = "/edit-excel?" + urlencode(edit_params)
            edited_badge = ' <span class="ok">✓ 已補</span>' if saved in edited else ""
            parts.append(
                f"<code>{html.escape(Path(saved).name)}</code> "
                f'<a class="button secondary small-button" href="{html.escape(edit_link, quote=True)}">補活動資訊</a>'
                f"{edited_badge}<br>"
            )
        if len(success) > 20:
            parts.append(f"...另有 {len(success) - 20} 個成功檔案，請到月份資料夾開啟。<br>")
    folder_link = "/open-folder?" + urlencode({"path": str(output_dir)})
    result_link = "/batch-result?" + urlencode({"job": job_id}) if job_id else ""
    parts.append(
        "<br>"
        f'<a class="button" href="{html.escape(folder_link, quote=True)}">開啟月份 EXCEL 資料夾</a> '
        + (f'<a class="button secondary" href="{html.escape(result_link, quote=True)}">回這次轉檔清單</a> ' if result_link else "")
        + f'<a class="button secondary" href="/batch-convert">回批次轉檔</a>'
    )
    return "".join(parts)


def render_batch_result_page(job_id):
    with BATCH_JOBS_LOCK:
        job = dict(BATCH_JOBS.get(job_id) or {})
    if not job:
        return render_batch_convert_page(error="找不到這次批次轉檔清單，可能是程式已重新啟動。")
    if job.get("status") != "done":
        return render_batch_progress_page(job_id)
    return f"""<!doctype html>
<html lang="zh-Hant">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>批次轉檔結果</title>
  <style>
    {base_styles()}
  </style>
</head>
<body>
  <main>
    {product_banner("批次轉檔結果")}
    {nav("batch")}
    <section class="status ok">
      {batch_convert_result_html(job, job_id)}
    </section>
  </main>
</body>
</html>"""


def batch_job_snapshot(job_id):
    with BATCH_JOBS_LOCK:
        job = dict(BATCH_JOBS.get(job_id) or {})
    if not job:
        return {"found": False}
    current = int(job.get("current", 0) or 0)
    total = int(job.get("total", 0) or 0)
    percent = round(current / total * 100) if total else 0
    snapshot = {
        "found": True,
        "status": job.get("status", "running"),
        "message": job.get("message", ""),
        "current": current,
        "total": total,
        "percent": percent,
        "error": job.get("error", ""),
        "result_html": "",
    }
    if snapshot["status"] == "done":
        snapshot["result_html"] = batch_convert_result_html(job, job_id)
    return snapshot


def run_batch_convert_job(job_id, month, metadata, fetch_weather, overwrite_newer, overwrite_all):
    rows = batch_scan(month)
    targets = [row for row in rows if should_batch_convert(row, overwrite_newer, overwrite_all)]
    result = {"success": [], "skipped": [row["fit"] for row in rows if row not in targets], "failed": []}
    total = len(targets)
    update_batch_job(job_id, total=total, current=0, message=f"找到 {total} 個需要轉檔的 FIT。")

    for index, row in enumerate(targets, start=1):
        update_batch_job(job_id, current=index - 1, message=f"轉檔中：{row['fit']}")
        try:
            saved = create_workbook(
                row["fit_path"],
                row["output_path"],
                metadata=metadata,
                fetch_weather=fetch_weather,
            )
            write_fit_to_sqlite(
                row["fit_path"],
                SQLITE_DB_PATH,
                metadata=metadata,
                fetch_weather=fetch_weather,
            )
            result["success"].append(str(saved))
            update_batch_job(job_id, current=index, message=f"已完成：{row['fit']}")
        except Exception as error:
            result["failed"].append({"fit": row["fit"], "error": friendly_error(error)})
            update_batch_job(job_id, current=index, message=f"轉檔失敗：{row['fit']}")

    update_batch_job(
        job_id,
        status="done",
        current=total,
        total=total,
        result=result,
        message="批次轉檔完成。",
    )


def start_batch_convert_job(month, metadata, fetch_weather, overwrite_newer, overwrite_all):
    job_id = uuid.uuid4().hex
    with BATCH_JOBS_LOCK:
        BATCH_JOBS[job_id] = {
            "status": "running",
            "message": "準備批次轉檔...",
            "current": 0,
            "total": 0,
            "month": month,
            "edited": [],
        }
    worker = threading.Thread(
        target=run_batch_convert_job,
        args=(job_id, month, metadata, fetch_weather, overwrite_newer, overwrite_all),
        daemon=True,
    )
    worker.start()
    return job_id


def parse_number(value):
    value = (value or "").strip()
    if not value:
        return ""
    try:
        return float(value)
    except ValueError:
        return value


def first_value(form, key, default=""):
    return (form.get(key, [default])[0] or "").strip()


def selected_values(form, key):
    return [value.strip() for value in form.get(key, []) if value.strip()]


def content_disposition_params(value):
    params = {}
    for part in value.split(";"):
        part = part.strip()
        if "=" not in part:
            continue
        key, raw = part.split("=", 1)
        params[key.lower()] = raw.strip().strip('"')
    return params


def parse_multipart(body, content_type):
    marker = "boundary="
    if marker not in content_type:
        return {}, {}
    boundary = content_type.split(marker, 1)[1].split(";", 1)[0].strip().strip('"').encode()
    delimiter = b"--" + boundary
    form = {}
    files = {}

    for part in body.split(delimiter):
        if not part or part in (b"--\r\n", b"--"):
            continue
        if part.startswith(b"\r\n"):
            part = part[2:]
        if part.endswith(b"--\r\n"):
            part = part[:-4]
        elif part.endswith(b"--"):
            part = part[:-2]
        if part.endswith(b"\r\n"):
            part = part[:-2]
        if b"\r\n\r\n" not in part:
            continue

        raw_headers, data = part.split(b"\r\n\r\n", 1)
        headers = {}
        for line in raw_headers.decode("utf-8", "replace").split("\r\n"):
            if ":" in line:
                key, value = line.split(":", 1)
                headers[key.lower()] = value.strip()
        disposition = headers.get("content-disposition", "")
        params = content_disposition_params(disposition)
        name = params.get("name")
        if not name:
            continue
        filename = params.get("filename")
        if filename:
            files[name] = {"filename": Path(filename).name, "content": data}
        else:
            form.setdefault(name, []).append(data.decode("utf-8", "replace"))
    return form, files


def parse_post_data(headers, body):
    content_type = headers.get("Content-Type", "")
    if content_type.startswith("multipart/form-data"):
        return parse_multipart(body, content_type)
    return parse_qs(body.decode("utf-8")), {}


def save_uploaded_fit(upload):
    if not upload:
        return None
    filename = upload.get("filename") or ""
    content = upload.get("content") or b""
    if not filename or not content:
        return None
    if Path(filename).suffix.lower() != ".fit":
        raise ValueError("上傳檔案必須是 .fit。")
    FIT_DIR.mkdir(parents=True, exist_ok=True)
    target = FIT_DIR / Path(filename).name
    target.write_bytes(content)
    return target


def build_metadata(form):
    return {
        "activity_name": first_value(form, "activity_name"),
        "shoe": first_value(form, "shoe"),
        "weather_temp": parse_number(first_value(form, "weather_temp")),
        "humidity": parse_number(first_value(form, "humidity")),
        "wind_direction": first_value(form, "wind_direction"),
        "wind_speed": first_value(form, "wind_speed"),
        "weather_description": first_value(form, "weather_description"),
        "workout_type": first_value(form, "workout_type"),
        "training_focus": "、".join(selected_values(form, "training_focus")),
        "feel": first_value(form, "feel"),
        "rpe": first_value(form, "rpe"),
        "fueling": first_value(form, "fueling"),
        "max_hr": parse_number(first_value(form, "max_hr")),
        "critical_power": parse_number(first_value(form, "critical_power")),
        "training_effect_aerobic": parse_number(first_value(form, "training_effect_aerobic")),
        "training_effect_anaerobic": parse_number(first_value(form, "training_effect_anaerobic")),
        "training_load": parse_number(first_value(form, "training_load")),
        "recovery_time_hr": parse_number(first_value(form, "recovery_time_hr")),
        "notes": first_value(form, "notes"),
    }


def option_tags(options, selected=""):
    tags = ['<option value="">自動 / 留空</option>']
    for option in options:
        value = html.escape(str(option), quote=True)
        is_selected = " selected" if str(option) == selected else ""
        tags.append(f'<option value="{value}"{is_selected}>{html.escape(str(option))}</option>')
    return "\n".join(tags)


def multi_option_tags(options, selected=None):
    selected = set(selected or [])
    tags = []
    for option in options:
        value = html.escape(str(option), quote=True)
        is_selected = " selected" if str(option) in selected else ""
        tags.append(f'<option value="{value}"{is_selected}>{html.escape(str(option))}</option>')
    return "\n".join(tags)


def options_with_existing(options, selected):
    result = list(options)
    for value in selected:
        if value not in result:
            result.append(value)
    return result


def activity_info_form_fields(values, dropdown_options):
    fields = []
    select_options = {
        "shoe": dropdown_options["shoes"],
        "workout_type": dropdown_options["workout_types"],
        "feel": dropdown_options["garmin_feel"],
        "rpe": dropdown_options["garmin_rpe"],
    }
    for key, _label, field_type, display in ACTIVITY_INFO_FIELDS:
        value = values.get(key, "")
        if field_type == "select":
            options = options_with_existing(select_options.get(key, []), [str(value)] if value not in ("", None) else [])
            fields.append(
                f"""
                <label>
                  <span>{html.escape(display)}</span>
                  <select name="{html.escape(key)}">{option_tags(options, str(value))}</select>
                </label>
                """
            )
        elif field_type == "multi":
            selected = [part.strip() for part in str(value or "").split("、") if part.strip()]
            options = options_with_existing(dropdown_options["training_focus"], selected)
            fields.append(
                f"""
                <label>
                  <span>{html.escape(display)}</span>
                  <select name="{html.escape(key)}" multiple>{multi_option_tags(options, selected)}</select>
                </label>
                """
            )
        elif field_type == "textarea":
            fields.append(
                f"""
                <label class="wide">
                  <span>{html.escape(display)}</span>
                  <textarea name="{html.escape(key)}">{html.escape(str(value or ""))}</textarea>
                </label>
                """
            )
        else:
            fields.append(input_field(display, key, value=value, input_type=field_type))
    return "\n".join(fields)


def input_field(label, name, value="", input_type="text", placeholder=""):
    return f"""
      <label>
        <span>{html.escape(label)}</span>
        <input type="{input_type}" name="{html.escape(name)}" value="{html.escape(str(value or ''), quote=True)}" placeholder="{html.escape(placeholder, quote=True)}">
      </label>
    """


def workout_focus_reference_table(dropdown_options):
    mapping = dropdown_options.get(WORKOUT_FOCUS_MAP_KEY, {})
    rows = []
    for workout in dropdown_options["workout_types"]:
        focus_matches = mapping.get(workout, [])
        focus_label = "、".join(focus_matches) if focus_matches else "未設定"
        focus_data = "||".join(focus_matches)
        rows.append(
            f"""
            <tr>
              <td>{html.escape(workout)}</td>
              <td>{html.escape(focus_label)}</td>
              <td>
                <button class="small-button" type="button" data-workout-value="{html.escape(workout, quote=True)}" data-focus-values="{html.escape(focus_data, quote=True)}">套用</button>
              </td>
            </tr>
            """
        )
    return f"""
      <section class="reference">
        <div class="reference-head">
          <h2>課表與訓練目的對照</h2>
        </div>
        <div class="reference-table-wrap">
          <table class="reference-table">
            <thead>
              <tr>
                <th>課表類型</th>
                <th>預設訓練目的</th>
                <th></th>
              </tr>
            </thead>
            <tbody>
              {"".join(rows)}
            </tbody>
          </table>
        </div>
      </section>
    """


def product_banner(title="FIT 匯入與資料整理"):
    return f"""
      <section class="product-banner">
        <div class="banner-copy">
          <div class="brand-row">
            <img class="brand-mark" src="/assets/rac_mark_transparent.png" alt="資料匯入工具">
            <div>
              <h1>{html.escape(title)}</h1>
              <p class="brand-subtitle">FIT IMPORT & DATA PREP</p>
            </div>
          </div>
          <p class="banner-subtitle">Garmin FIT -> Standardized Excel -> AI Coach Analysis</p>
          <div class="flow">
            <span>FIT Import</span>
            <span>{html.escape(EXCEL_SCHEMA_LABEL)}</span>
            <span>AI Coach</span>
            <span>Long-term Analytics</span>
          </div>
          <p class="banner-note">將 Garmin FIT 活動檔轉換為固定格式 Excel，支援每日 AI 教練分析、週/月趨勢與長期跑步資料庫。</p>
        </div>
      </section>
    """


def nav(active="convert"):
    convert_class = " active" if active == "convert" else ""
    download_class = " active" if active == "download-fit" else ""
    batch_class = " active" if active == "batch" else ""
    options_class = " active" if active == "options" else ""
    return f"""
      <nav>
        <a class="nav-link{convert_class}" href="/">轉檔</a>
        <a class="nav-link{download_class}" href="/download-fit">下載 FIT</a>
        <a class="nav-link{batch_class}" href="/batch-convert">批次轉檔</a>
        <a class="nav-link{options_class}" href="/options">下拉選單設定</a>
        <a class="nav-link nav-link-secondary" href="{html.escape(PLATFORM_URL, quote=True)}">回分析平台</a>
      </nav>
    """


def status_html(message="", error=""):
    if message:
        return f'<section class="status ok">{message}</section>'
    if error:
        return f'<section class="status error">{html.escape(error)}</section>'
    return ""


def base_styles():
    return """
    :root {
      color-scheme: light;
      --ink: #18222f;
      --muted: #657386;
      --line: #d9e3ee;
      --accent: #0f766e;
      --accent-dark: #0b4f5f;
      --accent-soft: #e7f4f2;
      --surface: #ffffff;
      --page: #f3f7fa;
      --error: #b42318;
      --ok: #166534;
    }
    * { box-sizing: border-box; }
    body {
      margin: 0;
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "Noto Sans TC", sans-serif;
      background:
        linear-gradient(180deg, #edf5f7 0, var(--page) 260px),
        var(--page);
      color: var(--ink);
    }
    main {
      width: min(1120px, calc(100vw - 32px));
      margin: 28px auto 40px;
    }
    h1 {
      margin: 0 0 6px;
      font-size: 30px;
      letter-spacing: 0;
    }
    .product-banner {
      display: grid;
      grid-template-columns: minmax(0, 1fr) auto;
      gap: 24px;
      align-items: start;
      margin: 0 0 18px;
      padding: 28px 32px;
      border-radius: 18px;
      color: #fff;
      background:
        linear-gradient(90deg, rgba(3, 33, 48, 0.7) 0%, rgba(5, 55, 65, 0.38) 46%, rgba(5, 87, 88, 0.2) 100%),
        url("/assets/rac_banner.png") center / cover no-repeat;
      box-shadow: 0 18px 48px rgba(11, 79, 95, 0.22);
      min-height: 286px;
    }
    .eyebrow {
      margin: 0 0 8px;
      font-size: 13px;
      font-weight: 800;
      letter-spacing: 0.04em;
      text-transform: uppercase;
      color: rgba(255, 255, 255, 0.76);
    }
    .product-banner h1 {
      margin: 0;
      font-size: 42px;
      line-height: 1.15;
      text-shadow: 0 2px 12px rgba(0, 0, 0, 0.24);
    }
    .brand-row {
      display: flex;
      align-items: center;
      gap: 24px;
    }
    .brand-mark {
      width: min(320px, 32vw);
      height: auto;
      filter: drop-shadow(0 10px 20px rgba(0, 0, 0, 0.28));
    }
    .brand-subtitle {
      margin: 8px 0 0;
      color: rgba(220, 236, 242, 0.76);
      font-size: 20px;
      font-weight: 800;
      letter-spacing: 0.08em;
    }
    .banner-subtitle {
      margin: 18px 0 0;
      font-size: 16px;
      color: rgba(255, 255, 255, 0.86);
    }
    .banner-note {
      max-width: 720px;
      margin: 16px 0 0;
      color: rgba(255, 255, 255, 0.86);
      line-height: 1.6;
    }
    .subtitle {
      margin: 0 0 18px;
      color: var(--muted);
      font-size: 15px;
      line-height: 1.6;
    }
    .flow {
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      margin: 18px 0 0;
    }
    .flow span {
      color: #fff;
      background: rgba(255, 255, 255, 0.14);
      border: 1px solid rgba(255, 255, 255, 0.2);
      border-radius: 999px;
      padding: 7px 10px;
      font-size: 13px;
      font-weight: 700;
    }
    .flow span + span::before {
      content: "-> ";
      opacity: 0.68;
    }
    nav {
      display: flex;
      width: fit-content;
      gap: 4px;
      margin: 0 0 18px;
      padding: 4px;
      border: 1px solid var(--line);
      border-radius: 999px;
      background: rgba(255, 255, 255, 0.74);
      box-shadow: 0 8px 24px rgba(31, 41, 51, 0.06);
    }
    .nav-link {
      color: var(--muted);
      text-decoration: none;
      padding: 9px 14px;
      border-radius: 999px;
      font-weight: 700;
    }
    .nav-link.active {
      color: #fff;
      background: var(--accent-dark);
    }
    .nav-link-secondary {
      color: var(--accent-dark);
      background: rgba(15, 118, 110, 0.08);
    }
    form {
      background: var(--surface);
      border: 1px solid var(--line);
      border-radius: 12px;
      padding: 24px;
      box-shadow: 0 16px 42px rgba(31, 41, 51, 0.1);
    }
    fieldset {
      border: 0;
      padding: 0;
      margin: 0 0 24px;
    }
    legend {
      padding: 0;
      margin: 0 0 14px;
      font-size: 17px;
      font-weight: 700;
    }
    .grid {
      display: grid;
      grid-template-columns: repeat(3, minmax(0, 1fr));
      gap: 14px;
    }
    label {
      display: flex;
      flex-direction: column;
      gap: 6px;
      min-width: 0;
    }
    label.wide { grid-column: span 3; }
    span {
      font-size: 13px;
      color: var(--muted);
    }
    input, select, textarea {
      width: 100%;
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 10px 11px;
      font: inherit;
      color: var(--ink);
      background: #fff;
    }
    input:focus, select:focus, textarea:focus {
      outline: 3px solid rgba(15, 118, 110, 0.16);
      border-color: var(--accent);
    }
    select[multiple] {
      min-height: 132px;
    }
    textarea {
      min-height: 76px;
      resize: vertical;
    }
    textarea.tall {
      min-height: 220px;
      line-height: 1.45;
      font-family: ui-monospace, SFMono-Regular, Menlo, Consolas, monospace;
      font-size: 13px;
    }
    .inline {
      display: flex;
      align-items: center;
      gap: 10px;
      color: var(--ink);
      margin-top: 2px;
    }
    .inline input {
      width: auto;
    }
    .actions {
      display: flex;
      align-items: center;
      gap: 12px;
      margin-top: 4px;
    }
    button, .button {
      appearance: none;
      border: 0;
      border-radius: 8px;
      background: var(--accent-dark);
      color: #fff;
      padding: 11px 16px;
      font: inherit;
      font-weight: 700;
      cursor: pointer;
      text-decoration: none;
      display: inline-block;
    }
    button:hover, .button:hover {
      background: var(--accent);
    }
    .secondary {
      background: #e7f0f3;
      color: var(--ink);
    }
    .secondary:hover {
      background: #d7e3ee;
    }
    .status {
      border-radius: 8px;
      padding: 14px 16px;
      margin: 0 0 18px;
      border: 1px solid var(--line);
      background: #fff;
      line-height: 1.55;
    }
    .ok { color: var(--ok); }
    .error { color: var(--error); }
    .note {
      color: var(--muted);
      font-size: 13px;
      margin: 8px 0 0;
    }
    .summary {
      width: 100%;
      border-collapse: collapse;
      margin: 10px 0 14px;
      color: var(--ink);
    }
    .summary th,
    .summary td {
      border-bottom: 1px solid var(--line);
      padding: 8px 6px;
      text-align: left;
      vertical-align: top;
    }
    .summary th {
      width: 130px;
      color: var(--muted);
      font-weight: 700;
    }
    .progress-card {
      background: var(--surface);
      border: 1px solid var(--line);
      border-radius: 12px;
      padding: 24px;
      box-shadow: 0 16px 42px rgba(31, 41, 51, 0.1);
    }
    .progress-bar {
      width: 100%;
      height: 14px;
      overflow: hidden;
      border-radius: 999px;
      background: #dbe7ee;
      margin: 14px 0 10px;
    }
    .progress-fill {
      width: 0%;
      height: 100%;
      border-radius: inherit;
      background: var(--accent-dark);
      transition: width 0.25s ease;
    }
    .progress-meta {
      display: flex;
      justify-content: space-between;
      gap: 12px;
      color: var(--muted);
      font-size: 13px;
    }
    .progress-message {
      margin: 0;
      color: var(--ink);
      font-weight: 700;
      line-height: 1.5;
    }
    .progress-result {
      margin-top: 18px;
      line-height: 1.65;
    }
    .reference {
      margin: 0 0 24px;
    }
    .reference-head {
      display: flex;
      align-items: baseline;
      justify-content: space-between;
      gap: 12px;
      margin: 0 0 10px;
    }
    .reference h2 {
      margin: 0;
      font-size: 17px;
      letter-spacing: 0;
    }
    .reference-table-wrap {
      overflow-x: auto;
      border: 1px solid var(--line);
      border-radius: 12px;
      background: #fff;
      box-shadow: 0 8px 22px rgba(31, 41, 51, 0.06);
    }
    .reference-table {
      width: 100%;
      min-width: 640px;
      border-collapse: collapse;
      color: var(--ink);
    }
    .reference-table th,
    .reference-table td {
      border-bottom: 1px solid var(--line);
      padding: 10px 12px;
      text-align: left;
      vertical-align: middle;
    }
    .reference-table th {
      background: #f7fafb;
      font-weight: 800;
    }
    .reference-table tr:last-child td {
      border-bottom: 0;
    }
    .small-button {
      padding: 7px 10px;
      font-size: 13px;
      white-space: nowrap;
    }
    code {
      font-family: ui-monospace, SFMono-Regular, Menlo, Consolas, monospace;
      font-size: 13px;
      color: var(--ink);
    }
    @media (max-width: 760px) {
      main { width: min(100vw - 20px, 1040px); margin: 18px auto; }
      .product-banner {
        grid-template-columns: 1fr;
        padding: 22px;
        border-radius: 14px;
        min-height: 0;
      }
      .brand-row {
        display: grid;
        gap: 12px;
      }
      .brand-mark {
        width: min(260px, 78vw);
      }
      .product-banner h1 { font-size: 30px; }
      .brand-subtitle { font-size: 14px; }
      form { padding: 16px; }
      .grid { grid-template-columns: 1fr; }
      label.wide { grid-column: span 1; }
      .actions { flex-direction: column; align-items: stretch; }
      button, .button { text-align: center; }
      nav { overflow-x: auto; }
      .nav-link { white-space: nowrap; }
    }
    """


def render_page(message="", error="", selected_fit=""):
    dropdown_options = load_app_options()
    files, total_count, list_limit = fit_files(selected_fit)
    fit_options = []
    for path in files:
        relative_name = relative_fit_path(path)
        value = html.escape(relative_name, quote=True)
        is_selected = " selected" if relative_name == selected_fit else ""
        fit_options.append(f'<option value="{value}"{is_selected}>{html.escape(relative_name)}</option>')
    if not fit_options:
        fit_options.append('<option value="">FIT 資料夾目前沒有 .fit 檔</option>')

    list_note = f"FIT 資料夾共有 {total_count} 個檔案，目前清單只顯示最近 {min(total_count, list_limit)} 個。"

    return f"""<!doctype html>
<html lang="zh-Hant">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>FIT 匯入與資料整理</title>
  <style>
    {base_styles()}
  </style>
</head>
<body>
  <main>
    {product_banner()}
    {nav("convert")}
    {status_html(message, error)}
    <form method="post" action="/convert" enctype="multipart/form-data">
      <fieldset>
        <legend>檔案</legend>
        <div class="grid">
          <label class="wide">
            <span>從電腦選擇 FIT 檔</span>
            <input type="file" name="upload_fit" accept=".fit">
          </label>
          <label class="wide">
            <span>或使用 FIT 資料夾裡的檔案</span>
            <select name="fit_file">
              {"".join(fit_options)}
            </select>
            <p class="note">{html.escape(list_note)}</p>
          </label>
          <label class="wide">
            <span>輸出檔名，可留空</span>
            <input name="output_name" placeholder="{html.escape(WORKBOOK_VERSION_NAME)}_YYYYMMDD_活動ID.xlsx">
          </label>
          <label class="inline wide">
            <input type="checkbox" name="fetch_weather" value="1" checked>
            <span>自動抓 Open-Meteo 歷史天氣</span>
          </label>
        </div>
      </fieldset>

      <fieldset>
        <legend>活動資訊</legend>
        <div class="grid">
          {input_field("活動名稱", "activity_name", placeholder="可留空")}
          <label>
            <span>鞋款</span>
            <select name="shoe">{option_tags(dropdown_options["shoes"])}</select>
          </label>
          <label>
            <span>課表類型</span>
            <select name="workout_type">{option_tags(dropdown_options["workout_types"])}</select>
          </label>
          <label>
            <span>訓練目的</span>
            <select name="training_focus" multiple>{multi_option_tags(dropdown_options["training_focus"])}</select>
            <p class="note">可多選；macOS 按 Command，Windows 按 Ctrl。</p>
          </label>
          <label>
            <span>感覺如何</span>
            <select name="feel">{option_tags(dropdown_options["garmin_feel"])}</select>
          </label>
          <label>
            <span>感受難度</span>
            <select name="rpe">{option_tags(dropdown_options["garmin_rpe"])}</select>
          </label>
          {input_field("最大心率", "max_hr", input_type="number", placeholder="自動")}
          {input_field("Critical Power(W)", "critical_power", input_type="number", placeholder="自動")}
          {input_field("氣溫(°C)", "weather_temp", input_type="number", placeholder="自動")}
          {input_field("濕度(%)", "humidity", input_type="number", placeholder="自動")}
          {input_field("風向", "wind_direction", placeholder="自動")}
          {input_field("風速", "wind_speed", placeholder="自動")}
          {input_field("天氣描述", "weather_description", placeholder="自動")}
          {input_field("Recovery Time (hr)", "recovery_time_hr", input_type="number")}
          {input_field("Training Load", "training_load", input_type="number", placeholder="自動")}
          <label class="wide">
            <span>補給紀錄</span>
            <textarea name="fueling"></textarea>
          </label>
          <label class="wide">
            <span>備註</span>
            <textarea name="notes"></textarea>
          </label>
        </div>
      </fieldset>

      {workout_focus_reference_table(dropdown_options)}

      <div class="actions">
        <button type="submit">轉成 Excel</button>
        <a class="button secondary" href="/">重新整理</a>
      </div>
    </form>
  </main>
  <script>
    function chooseOption(select, optionValue, allowMultiple) {{
      if (!select || !optionValue) return;
      for (const option of select.options) {{
        if (option.value === optionValue) {{
          if (allowMultiple) {{
            option.selected = true;
          }} else {{
            select.value = option.value;
          }}
          return;
        }}
      }}
    }}

    document.querySelectorAll("[data-workout-value]").forEach((button) => {{
      button.addEventListener("click", () => {{
        const workoutSelect = document.querySelector('select[name="workout_type"]');
        const focusSelect = document.querySelector('select[name="training_focus"]');
        chooseOption(workoutSelect, button.dataset.workoutValue, false);
        if (focusSelect) {{
          for (const option of focusSelect.options) option.selected = false;
          button.dataset.focusValues.split("||").forEach((focus) => chooseOption(focusSelect, focus, true));
        }}
      }});
    }});
  </script>
</body>
</html>"""


def year_options(selected_year):
    years = range(2020, dt.date.today().year + 2)
    tags = []
    for year in years:
        selected = " selected" if year == selected_year else ""
        tags.append(f'<option value="{year}"{selected}>{year}</option>')
    return "\n".join(tags)


def month_options(selected_month):
    tags = []
    for month in range(1, 13):
        selected = " selected" if month == selected_month else ""
        tags.append(f'<option value="{month}"{selected}>{month:02d}</option>')
    return "\n".join(tags)


def render_download_fit_page(message="", error=""):
    config = load_garmin_config()
    today = dt.date.today()
    remembered = " checked" if config.get("password") else ""
    return f"""<!doctype html>
<html lang="zh-Hant">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>下載 Garmin FIT</title>
  <style>
    {base_styles()}
  </style>
</head>
<body>
  <main>
    {product_banner("下載 Garmin FIT")}
    <p class="subtitle">從 Garmin Connect 下載指定月份的跑步活動 Original FIT，檔案會直接放進本專案的 FIT 資料夾，完成後可回到轉檔頁使用。</p>
    {nav("download-fit")}
    {status_html(message, error)}
    <form method="post" action="/download-fit">
      <fieldset>
        <legend>Garmin Connect</legend>
        <div class="grid">
          {input_field("電子信箱", "email", value=config.get("email", ""), input_type="email")}
          {input_field("密碼", "password", value=config.get("password", ""), input_type="password")}
          <label class="inline">
            <input type="checkbox" name="remember_garmin" value="1"{remembered}>
            <span>在本機記住 Garmin 登入設定</span>
          </label>
        </div>
      </fieldset>
      <fieldset>
        <legend>下載範圍</legend>
        <div class="grid">
          <label class="inline wide">
            <input type="radio" name="download_mode" value="today">
            <span>只下載今天</span>
          </label>
          <label class="inline wide">
            <input type="radio" name="download_mode" value="month" checked>
            <span>下載指定月份</span>
          </label>
          <label>
            <span>年份</span>
            <select name="year" data-month-field>{year_options(today.year)}</select>
          </label>
          <label>
            <span>月份</span>
            <select name="month" data-month-field>{month_options(today.month)}</select>
          </label>
          <label>
            <span>儲存位置</span>
            <input value="{html.escape(str(FIT_DIR / 'YYYY-MM'), quote=True)}" disabled>
          </label>
        </div>
      </fieldset>
      <div class="actions">
        <button type="submit">下載 FIT</button>
        <a class="button secondary" href="/">回轉檔</a>
      </div>
    </form>
  </main>
  <script>
    function updateDownloadMode() {{
      const todayMode = document.querySelector('input[name="download_mode"][value="today"]')?.checked;
      document.querySelectorAll("[data-month-field]").forEach((field) => {{
        field.disabled = todayMode;
      }});
    }}
    document.querySelectorAll('input[name="download_mode"]').forEach((radio) => {{
      radio.addEventListener("change", updateDownloadMode);
    }});
    updateDownloadMode();
  </script>
</body>
</html>"""


def render_download_progress_page(job_id):
    return f"""<!doctype html>
<html lang="zh-Hant">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>下載進度</title>
  <style>
    {base_styles()}
  </style>
</head>
<body>
  <main>
    {product_banner("下載 Garmin FIT")}
    {nav("download-fit")}
    <section class="progress-card">
      <p id="progress-message" class="progress-message">準備下載...</p>
      <div class="progress-bar" aria-label="下載進度">
        <div id="progress-fill" class="progress-fill"></div>
      </div>
      <div class="progress-meta">
        <span id="progress-count">0 / 0</span>
        <span id="progress-percent">0%</span>
      </div>
      <div id="progress-result" class="progress-result"></div>
    </section>
  </main>
  <script>
    const jobId = "{html.escape(job_id, quote=True)}";
    async function refreshProgress() {{
      const response = await fetch(`/download-fit-status?job=${{encodeURIComponent(jobId)}}`, {{cache: "no-store"}});
      const data = await response.json();
      const message = document.getElementById("progress-message");
      const fill = document.getElementById("progress-fill");
      const count = document.getElementById("progress-count");
      const percent = document.getElementById("progress-percent");
      const result = document.getElementById("progress-result");

      if (!data.found) {{
        message.textContent = "找不到下載工作，請重新開始。";
        result.innerHTML = '<a class="button secondary" href="/download-fit">回下載頁</a>';
        return;
      }}

      message.textContent = data.error || data.message || "下載中...";
      fill.style.width = `${{data.percent || 0}}%`;
      count.textContent = `${{data.current || 0}} / ${{data.total || 0}}`;
      percent.textContent = `${{data.percent || 0}}%`;

      if (data.status === "done") {{
        fill.style.width = "100%";
        percent.textContent = "100%";
        result.innerHTML = data.result_html || "";
        return;
      }}
      if (data.status === "error") {{
        result.innerHTML = '<a class="button secondary" href="/download-fit">重新下載</a>';
        return;
      }}
      window.setTimeout(refreshProgress, 1000);
    }}
    refreshProgress();
  </script>
</body>
</html>"""


def month_select_options(selected_month):
    months = available_fit_months()
    if not months:
        today = dt.date.today().strftime("%Y-%m")
        months = [today]
    tags = []
    for month in months:
        selected = " selected" if month == selected_month else ""
        tags.append(f'<option value="{html.escape(month, quote=True)}"{selected}>{html.escape(month)}</option>')
    return "\n".join(tags)


def render_batch_convert_page(message="", error="", selected_month=""):
    months = available_fit_months()
    selected_month = selected_month or (months[0] if months else dt.date.today().strftime("%Y-%m"))
    options = load_app_options()
    return f"""<!doctype html>
<html lang="zh-Hant">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>批次轉檔</title>
  <style>
    {base_styles()}
  </style>
</head>
<body>
  <main>
    {product_banner("批次轉檔")}
    <p class="subtitle">選擇月份後，系統會掃描本機 FIT 資料夾並轉換尚未產生 Excel 的活動。</p>
    {nav("batch")}
    {status_html(message, error)}
    <form method="post" action="/batch-convert">
      <fieldset>
        <legend>月份與共用設定</legend>
        <div class="grid">
          <label>
            <span>月份</span>
            <select name="month" id="batch-month">{month_select_options(selected_month)}</select>
          </label>
          <label>
            <span>鞋款</span>
            <select name="shoe">{option_tags(options["shoes"])}</select>
          </label>
          {input_field("最大心率", "max_hr", input_type="number", placeholder="自動")}
          {input_field("Critical Power(W)", "critical_power", input_type="number", placeholder="自動")}
          <label class="inline wide">
            <input type="checkbox" name="fetch_weather" value="1" checked>
            <span>自動抓 Open-Meteo 歷史天氣</span>
          </label>
          <label class="inline wide">
            <input type="checkbox" name="overwrite_newer" value="1" checked>
            <span>FIT 比 Excel 新時自動重轉</span>
          </label>
          <label class="inline wide">
            <input type="checkbox" name="overwrite_all" value="1">
            <span>全部重轉</span>
          </label>
        </div>
      </fieldset>
      <fieldset>
        <legend>目前狀態</legend>
        {batch_status_table(selected_month)}
      </fieldset>
      <div class="actions">
        <button type="submit">轉換未完成項目</button>
        <a class="button secondary" href="/batch-convert">重新掃描</a>
      </div>
    </form>
  </main>
  <script>
    document.getElementById("batch-month")?.addEventListener("change", (event) => {{
      window.location.href = `/batch-convert?month=${{encodeURIComponent(event.target.value)}}`;
    }});
  </script>
</body>
</html>"""


def render_batch_progress_page(job_id):
    return f"""<!doctype html>
<html lang="zh-Hant">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>批次轉檔進度</title>
  <style>
    {base_styles()}
  </style>
</head>
<body>
  <main>
    {product_banner("批次轉檔")}
    {nav("batch")}
    <section class="progress-card">
      <p id="progress-message" class="progress-message">準備批次轉檔...</p>
      <div class="progress-bar" aria-label="批次轉檔進度">
        <div id="progress-fill" class="progress-fill"></div>
      </div>
      <div class="progress-meta">
        <span id="progress-count">0 / 0</span>
        <span id="progress-percent">0%</span>
      </div>
      <div id="progress-result" class="progress-result"></div>
    </section>
  </main>
  <script>
    const jobId = "{html.escape(job_id, quote=True)}";
    async function refreshProgress() {{
      const response = await fetch(`/batch-convert-status?job=${{encodeURIComponent(jobId)}}`, {{cache: "no-store"}});
      const data = await response.json();
      const message = document.getElementById("progress-message");
      const fill = document.getElementById("progress-fill");
      const count = document.getElementById("progress-count");
      const percent = document.getElementById("progress-percent");
      const result = document.getElementById("progress-result");

      if (!data.found) {{
        message.textContent = "找不到批次轉檔工作，請重新開始。";
        result.innerHTML = '<a class="button secondary" href="/batch-convert">回批次轉檔</a>';
        return;
      }}
      message.textContent = data.error || data.message || "轉檔中...";
      fill.style.width = `${{data.percent || 0}}%`;
      count.textContent = `${{data.current || 0}} / ${{data.total || 0}}`;
      percent.textContent = `${{data.percent || 0}}%`;
      if (data.status === "done") {{
        fill.style.width = "100%";
        percent.textContent = "100%";
        result.innerHTML = data.result_html || "";
        return;
      }}
      if (data.status === "error") {{
        result.innerHTML = '<a class="button secondary" href="/batch-convert">重新開始</a>';
        return;
      }}
      window.setTimeout(refreshProgress, 1000);
    }}
    refreshProgress();
  </script>
</body>
</html>"""


def render_edit_excel_page(path, message="", error="", job_id=""):
    if not is_output_file(path):
        return render_page(error="找不到可編輯的 Excel 檔。")
    dropdown_options = load_app_options()
    try:
        values = load_activity_info_values(path)
        summary = summary_html(workbook_summary(path))
    except Exception as read_error:
        return render_page(error=f"讀取 Excel 失敗：{read_error}")

    download_link = "/download?" + urlencode({"path": str(path)})
    open_link = "/open?" + urlencode({"path": str(path)})
    folder_link = "/open-folder?" + urlencode({"path": str(path.parent)})
    back_link = "/batch-result?" + urlencode({"job": job_id}) if job_id else "/batch-convert"
    back_label = "回這次轉檔清單" if job_id else "回批次轉檔"
    return f"""<!doctype html>
<html lang="zh-Hant">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>補活動資訊</title>
  <style>
    {base_styles()}
  </style>
</head>
<body>
  <main>
    {product_banner("補活動資訊")}
    <p class="subtitle"><code>{html.escape(str(path))}</code></p>
    {nav("batch")}
    {status_html(message, error)}
    {summary}
    <form method="post" action="/edit-excel">
      <input type="hidden" name="path" value="{html.escape(str(path), quote=True)}">
      <input type="hidden" name="job" value="{html.escape(job_id, quote=True)}">
      <fieldset>
        <legend>活動資訊</legend>
        <div class="grid">
          {activity_info_form_fields(values, dropdown_options)}
        </div>
      </fieldset>
      <div class="actions">
        <button type="submit">儲存活動資訊</button>
        <a class="button secondary" href="{html.escape(back_link, quote=True)}">{html.escape(back_label)}</a>
        <a class="button secondary" href="{html.escape(open_link, quote=True)}">開啟 Excel</a>
        <a class="button secondary" href="{html.escape(download_link, quote=True)}">下載 Excel</a>
        <a class="button secondary" href="{html.escape(folder_link, quote=True)}">開啟資料夾</a>
      </div>
    </form>
  </main>
</body>
</html>"""


def options_textarea(name, label, values):
    text = "\n".join(str(value) for value in values)
    return f"""
      <label class="wide">
        <span>{html.escape(label)}，每行一個選項</span>
        <textarea class="tall" name="{html.escape(name)}">{html.escape(text)}</textarea>
      </label>
    """


def mapping_select(name, training_focus, selected):
    return f"""
      <select name="{html.escape(name)}" multiple>
        {multi_option_tags(training_focus, selected)}
      </select>
    """


def mapping_settings_table(options):
    rows = []
    mapping = options.get(WORKOUT_FOCUS_MAP_KEY, {})
    training_focus = options["training_focus"]
    for index, workout in enumerate(options["workout_types"]):
        selected = mapping.get(workout, [])
        rows.append(
            f"""
            <tr>
              <td>
                {html.escape(workout)}
                <input type="hidden" name="map_workout_{index}" value="{html.escape(workout, quote=True)}">
              </td>
              <td>{mapping_select(f"map_focus_{index}", training_focus, selected)}</td>
            </tr>
            """
        )
    return f"""
      <div class="reference-table-wrap">
        <table class="reference-table">
          <thead>
            <tr>
              <th>課表類型</th>
              <th>對應訓練目的</th>
            </tr>
          </thead>
          <tbody>
            {"".join(rows)}
          </tbody>
        </table>
      </div>
    """


def dropdown_options_from_form(form):
    result = {}
    for key, _label in OPTION_FIELDS:
        lines = [line.strip() for line in first_value(form, key).splitlines()]
        values = []
        seen = set()
        for line in lines:
            if not line or line in seen:
                continue
            values.append(line)
            seen.add(line)
        if not values:
            raise ValueError("每一組下拉選單至少需要一個選項。")
        result[key] = values
    result[WORKOUT_FOCUS_MAP_KEY] = workout_focus_map_from_form(form, result)
    return result


def workout_focus_map_from_form(form, options):
    valid_workouts = set(options["workout_types"])
    valid_focus = set(options["training_focus"])
    result = {}
    for index, workout in enumerate(options["workout_types"]):
        posted_workout = first_value(form, f"map_workout_{index}") or workout
        if posted_workout not in valid_workouts:
            continue
        selected = [
            value
            for value in selected_values(form, f"map_focus_{index}")
            if value in valid_focus
        ]
        result[posted_workout] = selected
    defaults = default_workout_focus_map(options)
    for workout in options["workout_types"]:
        result.setdefault(workout, defaults.get(workout, []))
    return result


def save_dropdown_options(options):
    CONFIG_DIR = DROPDOWN_CONFIG_PATH.parent
    CONFIG_DIR.mkdir(parents=True, exist_ok=True)
    DROPDOWN_CONFIG_PATH.write_text(
        json.dumps(options, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def render_options_page(message="", error=""):
    options = load_app_options()
    fields = "\n".join(options_textarea(key, label, options[key]) for key, label in OPTION_FIELDS)
    return f"""<!doctype html>
<html lang="zh-Hant">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>下拉選單設定</title>
  <style>
    {base_styles()}
  </style>
</head>
<body>
  <main>
    {product_banner("下拉選單設定")}
    <p class="subtitle">先修改活動資訊選項並儲存，再設定課表類型與訓練目的的對應關係。儲存後會立即套用到轉檔頁與輸出的 Excel。</p>
    {nav("options")}
    {status_html(message, error)}
    <form method="post" action="/options">
      <fieldset>
        <legend>選項內容</legend>
        <div class="grid">
          {fields}
        </div>
      </fieldset>
      <fieldset>
        <legend>課表與訓練目的對應</legend>
        <p class="note">若剛新增或改名課表/訓練目的，請先儲存選項內容；頁面重新整理後再設定對應關係。</p>
        {mapping_settings_table(options)}
      </fieldset>
      <div class="actions">
        <button type="submit">儲存設定</button>
        <a class="button secondary" href="/">回轉檔</a>
      </div>
    </form>
  </main>
</body>
</html>"""


class AppHandler(BaseHTTPRequestHandler):
    def send_html(self, content, status=200):
        data = content.encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def send_json(self, payload, status=200):
        data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(data)

    def send_xlsx(self, path):
        data = path.read_bytes()
        filename = path.name
        self.send_response(200)
        self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Length", str(len(data)))
        self.send_header(
            "Content-Disposition",
            f"attachment; filename*=UTF-8''{quote(filename)}",
        )
        self.end_headers()
        self.wfile.write(data)

    def send_asset(self, path):
        try:
            resolved = path.resolve()
            asset_root = ASSETS_DIR.resolve()
            if not resolved.is_file() or not resolved.is_relative_to(asset_root):
                raise FileNotFoundError
            data = resolved.read_bytes()
        except (OSError, FileNotFoundError):
            self.send_html(render_page(error="找不到指定的介面資產。"), status=404)
            return
        content_type = mimetypes.guess_type(str(resolved))[0] or "application/octet-stream"
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(data)))
        self.send_header("Cache-Control", "no-cache")
        self.end_headers()
        self.wfile.write(data)

    def do_GET(self):
        parsed = urlparse(self.path)
        query = parse_qs(parsed.query)
        if parsed.path.startswith("/assets/"):
            self.send_asset(ASSETS_DIR / parsed.path.removeprefix("/assets/"))
            return
        if parsed.path == "/options":
            self.send_html(render_options_page())
            return
        if parsed.path == "/download-fit":
            self.send_html(render_download_fit_page())
            return
        if parsed.path == "/download-fit-status":
            self.send_json(download_job_snapshot(first_value(query, "job")))
            return
        if parsed.path == "/batch-convert":
            self.send_html(render_batch_convert_page(selected_month=first_value(query, "month")))
            return
        if parsed.path == "/batch-convert-status":
            self.send_json(batch_job_snapshot(first_value(query, "job")))
            return
        if parsed.path == "/batch-result":
            self.send_html(render_batch_result_page(first_value(query, "job")))
            return
        if parsed.path == "/edit-excel":
            self.send_html(render_edit_excel_page(Path(first_value(query, "path")), job_id=first_value(query, "job")))
            return
        if parsed.path == "/open":
            output = Path(first_value(query, "path"))
            if is_output_file(output):
                open_file(output)
                self.send_html(render_page(message=f"已要求系統開啟 <code>{html.escape(str(output))}</code>"))
            else:
                self.send_html(render_page(error="找不到輸出檔。"), status=404)
            return
        if parsed.path == "/download":
            output = Path(first_value(query, "path"))
            if is_output_file(output):
                self.send_xlsx(output)
            else:
                self.send_html(render_page(error="找不到可下載的 Excel 檔。"), status=404)
            return
        if parsed.path == "/open-folder":
            folder = Path(first_value(query, "path")) if first_value(query, "path") else DEFAULT_OUTPUT_DIR
            if folder == DEFAULT_OUTPUT_DIR:
                folder.mkdir(parents=True, exist_ok=True)
            if is_output_dir(folder):
                open_file(folder)
                self.send_html(render_page(message=f"已要求系統開啟 <code>{html.escape(str(folder))}</code>"))
            else:
                self.send_html(render_page(error="找不到可開啟的 EXCEL 資料夾。"), status=404)
            return
        self.send_html(render_page())

    def do_POST(self):
        parsed = urlparse(self.path)
        if parsed.path == "/edit-excel":
            length = int(self.headers.get("Content-Length", "0"))
            body = self.rfile.read(length)
            form, _files = parse_post_data(self.headers, body)
            path = Path(first_value(form, "path"))
            job_id = first_value(form, "job")
            if not is_output_file(path):
                self.send_html(render_page(error="找不到可編輯的 Excel 檔。"), status=404)
                return
            try:
                save_activity_info_values(path, activity_info_values_from_form(form))
            except Exception as error:
                self.send_html(render_edit_excel_page(path, error=f"儲存失敗：{error}", job_id=job_id), status=500)
                return
            if job_id:
                with BATCH_JOBS_LOCK:
                    job = BATCH_JOBS.get(job_id)
                    if job is not None:
                        edited = job.setdefault("edited", [])
                        saved_path = str(path)
                        if saved_path not in edited:
                            edited.append(saved_path)
            self.send_html(render_edit_excel_page(path, message="活動資訊已儲存。", job_id=job_id))
            return

        if parsed.path == "/batch-convert":
            length = int(self.headers.get("Content-Length", "0"))
            body = self.rfile.read(length)
            form, _files = parse_post_data(self.headers, body)
            month = first_value(form, "month")
            if not re.fullmatch(r"\d{4}-\d{2}", month or ""):
                self.send_html(render_batch_convert_page(error="請選擇有效月份。"), status=400)
                return
            if not fit_files_for_month(month):
                self.send_html(render_batch_convert_page(error="這個月份沒有 FIT 檔。", selected_month=month), status=400)
                return
            metadata = {
                "shoe": first_value(form, "shoe"),
                "max_hr": parse_number(first_value(form, "max_hr")),
                "critical_power": parse_number(first_value(form, "critical_power")),
            }
            fetch_weather = first_value(form, "fetch_weather") == "1"
            overwrite_newer = first_value(form, "overwrite_newer") == "1"
            overwrite_all = first_value(form, "overwrite_all") == "1"
            job_id = start_batch_convert_job(month, metadata, fetch_weather, overwrite_newer, overwrite_all)
            self.send_html(render_batch_progress_page(job_id))
            return

        if parsed.path == "/download-fit":
            length = int(self.headers.get("Content-Length", "0"))
            body = self.rfile.read(length)
            form, _files = parse_post_data(self.headers, body)

            email = first_value(form, "email")
            password = first_value(form, "password")
            remember_garmin = first_value(form, "remember_garmin") == "1"
            download_mode = first_value(form, "download_mode") or "month"
            try:
                if download_mode == "today":
                    today = dt.date.today()
                    year = today.year
                    month = today.month
                else:
                    year = int(first_value(form, "year"))
                    month = int(first_value(form, "month"))
                    if not 1 <= month <= 12:
                        raise ValueError
            except ValueError:
                self.send_html(render_download_fit_page(error="請選擇有效的下載月份。"), status=400)
                return

            if not email or not password:
                self.send_html(render_download_fit_page(error="請輸入 Garmin 帳號與密碼。"), status=400)
                return

            if remember_garmin:
                save_garmin_config(email, password)
            else:
                remove_garmin_config()

            job_id = start_download_job(email, password, year, month, mode=download_mode)
            self.send_html(render_download_progress_page(job_id))
            return

        if parsed.path == "/options":
            length = int(self.headers.get("Content-Length", "0"))
            body = self.rfile.read(length)
            form, _files = parse_post_data(self.headers, body)
            try:
                options = dropdown_options_from_form(form)
                save_dropdown_options(options)
            except Exception as error:
                self.send_html(render_options_page(error=f"儲存失敗：{error}"), status=400)
                return
            self.send_html(render_options_page(message="下拉選單已更新。"))
            return

        if parsed.path != "/convert":
            self.send_html(render_page(error="不支援的操作。"), status=404)
            return

        length = int(self.headers.get("Content-Length", "0"))
        body = self.rfile.read(length)
        try:
            form, files = parse_post_data(self.headers, body)
            fit_path = save_uploaded_fit(files.get("upload_fit"))
        except ValueError as error:
            self.send_html(render_page(error=str(error)), status=400)
            return

        fit_name = first_value(form, "fit_file")
        if fit_path is None:
            fit_path = resolve_fit_path(fit_name)
        if not fit_path.exists() or fit_path.suffix.lower() != ".fit":
            self.send_html(render_page(error="請選擇一個 .fit 檔，或從 FIT 資料夾清單選擇有效檔案。"), status=400)
            return
        fit_name = relative_fit_path(fit_path)

        output_name = first_value(form, "output_name")
        output_name = Path(output_name).name if output_name else ""
        output_path = excel_output_dir_for_fit(fit_path) / output_name if output_name else excel_output_path_for_fit(fit_path)
        if output_path.suffix.lower() != ".xlsx":
            output_path = output_path.with_suffix(".xlsx")

        metadata = build_metadata(form)
        fetch_weather = first_value(form, "fetch_weather") == "1"

        try:
            saved = create_workbook(fit_path, output_path, metadata=metadata, fetch_weather=fetch_weather)
            write_fit_to_sqlite(fit_path, SQLITE_DB_PATH, metadata=metadata, fetch_weather=fetch_weather)
        except Exception as error:
            self.send_html(render_page(error=friendly_error(error), selected_fit=fit_name), status=500)
            return

        open_link = "/open?" + urlencode({"path": str(saved)})
        download_link = "/download?" + urlencode({"path": str(saved)})
        folder_link = "/open-folder?" + urlencode({"path": str(saved.parent)})
        try:
            summary = summary_html(workbook_summary(saved))
        except Exception:
            summary = ""
        message = (
            f"轉檔完成：<code>{html.escape(str(saved))}</code><br>"
            f"{summary}"
            f'<a class="button" href="{html.escape(open_link, quote=True)}">開啟 Excel</a> '
            f'<a class="button secondary" href="{html.escape(download_link, quote=True)}">下載 Excel</a> '
            f'<a class="button secondary" href="{html.escape(folder_link, quote=True)}">開啟 EXCEL 資料夾</a>'
        )
        self.send_html(render_page(message=message, selected_fit=fit_name))

    def log_message(self, format, *args):
        return


def open_browser_later(url):
    timer = threading.Timer(0.6, lambda: webbrowser.open(url))
    timer.daemon = True
    timer.start()


def main():
    try:
        server = ThreadingHTTPServer((HOST, PORT), AppHandler)
    except PermissionError:
        print(f"無法啟動本機網站：系統目前不允許使用 {HOST}:{PORT}。")
        print("請確認防火牆或安全性設定，或改用 CLI 方式轉檔。")
        return
    except OSError as error:
        if getattr(error, "errno", None) == 48:
            print(f"無法啟動本機網站：{HOST}:{PORT} 已經被其他程式使用。")
            print("請關掉舊的轉檔視窗，或稍後再重新啟動。")
        else:
            print(f"無法啟動本機網站：{error}")
        return
    url = f"http://{HOST}:{PORT}"
    print(f"Running Analytics v{APP_VERSION}: {url}")
    open_browser_later(url)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nStopped.")


if __name__ == "__main__":
    main()
