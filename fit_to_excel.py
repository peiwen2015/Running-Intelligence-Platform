#!/usr/bin/env python3
from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import re
import sqlite3
from pathlib import Path
from statistics import mean
from urllib.parse import urlencode
from urllib.request import urlopen

from garmin_fit_sdk import Decoder, Stream
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


FIT_EPOCH = 631065600
APP_VERSION = "1.0.0"
WORKBOOK_VERSION_NAME = "跑步分析資料 v1.1"
DEFAULT_OUTPUT_DIR = Path(__file__).resolve().parent / "EXCEL"
CONFIG_DIR = Path(__file__).resolve().parent / "config"
DROPDOWN_CONFIG_PATH = CONFIG_DIR / "dropdown_options.json"
SQLITE_DB_PATH = Path(__file__).resolve().parent / "analysis_platform" / "running_analytics.sqlite"
SQLITE_SCHEMA_PATH = Path(__file__).resolve().parent / "docs" / "30_Physical_Model" / "SQLite Schema v1.0.sql"
FEEDBACK_DICTIONARY_SEED_PATH = Path(__file__).resolve().parent / "analysis_platform" / "feedback_dictionary_seed.json"
STAMINA_RECORD_FIELDS = (137, 138)
STAMINA_SESSION_START = 205
STAMINA_SESSION_END_FIELDS = (206, 207)

DEFAULT_DROPDOWN_OPTIONS = {
    "shoes": [],
    "workout_types": [
    "Recovery Run（恢復跑）",
    "Easy Run（輕鬆跑）",
    "LSD（長距離慢跑）",
    "Long Run（長跑）",
    "Tempo Run（節奏跑）",
    "Marathon Pace（馬拉松配速）",
    "Interval（間歇）",
    "Repetition（速度訓練）",
    "Progression Run（漸速跑）",
    "Fartlek（法特萊克）",
    "Race（比賽）",
    "Other（其他）",
    ],
    "training_focus": [
    "Recovery",
    "Aerobic Base",
    "Endurance",
    "Marathon Pace",
    "Threshold",
    "VO₂max",
    "Speed",
    "Running Economy",
    "Heat Adaptation",
    "Race",
    ],
    "garmin_rpe": [
    "1 - 非常輕鬆",
    "2 - 輕鬆",
    "3 - 中等",
    "4 - 有點難",
    "5 - 困難",
    "6 - 困難",
    "7 - 非常困難",
    "8 - 非常困難",
    "9 - 超級難",
    "10 - 極限",
    ],
    "garmin_feel": [
    "非常弱",
    "弱",
    "普通",
    "強",
    "非常強",
    ],
}
WEATHER_FIELDS = ("weather_temp", "humidity", "wind_direction", "wind_speed", "weather_description")
DROPDOWN_SOURCE_TABLE = "dropdown_source"
FEEDBACK_DICTIONARY_TABLE = "feedback_dictionary_option"

SHOE_DIMENSION_DEFAULTS = {}

WORKOUT_TYPE_DIMENSION_DEFAULTS = {
    "Recovery Run": ("recovery_run", "Recovery Run", "恢復跑", "Recovery", 0, 0, 1, 10, "#7CB7B8"),
    "Easy Run": ("easy_run", "Easy Run", "輕鬆跑", "Easy", 0, 0, 0, 20, "#6FA8DC"),
    "LSD": ("lsd", "LSD", "長距離慢跑", "Easy", 0, 1, 0, 30, "#93C47D"),
    "Long Run": ("long_run", "Long Run", "長跑", "Moderate", 0, 1, 0, 40, "#76A5AF"),
    "Tempo Run": ("tempo_run", "Tempo Run", "節奏跑", "Quality", 1, 0, 0, 50, "#E69138"),
    "Marathon Pace": ("marathon_pace", "Marathon Pace", "馬拉松配速", "Quality", 1, 0, 0, 60, "#F6B26B"),
    "Interval": ("interval", "Interval", "間歇", "Quality", 1, 0, 0, 70, "#CC0000"),
    "Repetition": ("repetition", "Repetition", "速度訓練", "Quality", 1, 0, 0, 80, "#990000"),
    "Progression Run": ("progression_run", "Progression Run", "漸速跑", "Moderate", 1, 0, 0, 90, "#B6D7A8"),
    "Fartlek": ("fartlek", "Fartlek", "法特萊克", "Quality", 1, 0, 0, 100, "#674EA7"),
    "Race": ("race", "Race", "比賽", "Race", 1, 0, 0, 110, "#000000"),
    "Other": ("other", "Other", "其他", "Moderate", 0, 0, 0, 120, "#999999"),
}

TRAINING_PURPOSE_DIMENSION_DEFAULTS = {
    "Recovery": ("recovery", "Recovery", "恢復", "Recovery", 0, 1, 0, 10, "#7CB7B8"),
    "Aerobic Base": ("aerobic_base", "Aerobic Base", "有氧基礎", "Aerobic", 1, 0, 0, 20, "#6FA8DC"),
    "Endurance": ("endurance", "Endurance", "耐力", "Endurance", 1, 0, 1, 30, "#93C47D"),
    "Marathon Pace": ("race_specific", "Race Specific", "比賽專項", "Race", 1, 0, 1, 40, "#F6B26B"),
    "Threshold": ("threshold", "Threshold", "乳酸閾值", "Threshold", 1, 0, 1, 50, "#E69138"),
    "VO2max": ("vo2max", "VO2max", "最大攝氧量", "VO2max", 1, 0, 1, 60, "#CC0000"),
    "Speed": ("speed", "Speed", "速度", "Speed", 1, 0, 1, 70, "#990000"),
    "Neuromuscular": ("neuromuscular", "Neuromuscular", "神經肌肉活化", "Technique", 0, 0, 1, 80, "#674EA7"),
    "Running Economy": ("running_economy", "Running Economy", "跑步經濟性", "Technique", 0, 0, 1, 90, "#8E7CC3"),
    "Running Form": ("running_economy", "Running Economy", "跑步經濟性", "Technique", 0, 0, 1, 90, "#8E7CC3"),
    "Heat Adaptation": ("heat_adaptation", "Heat Adaptation", "高溫適應", "Environmental", 0, 0, 1, 100, "#F1C232"),
    "Race Simulation": ("race_specific", "Race Specific", "比賽專項", "Race", 1, 0, 1, 40, "#F6B26B"),
    "Race": ("race", "Race", "正式比賽", "Race", 0, 0, 1, 110, "#000000"),
    "Taper": ("maintenance", "Maintenance", "維持", "Maintenance", 0, 1, 0, 120, "#B7B7B7"),
    "Test": ("race", "Race", "正式比賽", "Race", 0, 0, 1, 110, "#000000"),
}


HEADERS = [
    "公里",
    "距離(m)",
    "時間(秒)",
    "配速(分:秒/km)",
    "平均心率",
    "平均心率%",
    "最高心率",
    "平均步頻(spm)",
    "平均功率(W)",
    "平均功率%",
    "垂直振幅(mm)",
    "垂直比(%)",
    "觸地時間(ms)",
    "步幅(mm)",
    "溫度(°C)",
    "Stamina 起",
    "Stamina 末",
    "爬升(m)",
]


def fit_datetime(value):
    if isinstance(value, dt.datetime):
        return value
    if isinstance(value, (int, float)):
        return dt.datetime.fromtimestamp(value + FIT_EPOCH, tz=dt.timezone.utc)
    return None


def pace_text(seconds: float, meters: float) -> str:
    if not meters:
        return ""
    sec_per_km = round(seconds / meters * 1000)
    return f"{sec_per_km // 60}:{sec_per_km % 60:02d}"


def rounded(value, ndigits=1):
    if value is None:
        return None
    return round(float(value), ndigits)


def null_if_blank(value):
    return None if value in ("", None) else value


def int_or_none(value):
    if value in ("", None):
        return None
    try:
        return int(round(float(value)))
    except (TypeError, ValueError):
        return None


def float_or_none(value):
    if value in ("", None):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def cell_value(value):
    return "" if value is None else value


def average(values, ndigits=1):
    vals = [float(v) for v in values if isinstance(v, (int, float))]
    return round(mean(vals), ndigits) if vals else None


def weighted_average(value_weight_pairs, ndigits=1):
    total_weight = 0.0
    weighted_sum = 0.0
    for value, weight in value_weight_pairs:
        if not isinstance(value, (int, float)) or not isinstance(weight, (int, float)) or weight <= 0:
            continue
        total_weight += float(weight)
        weighted_sum += float(value) * float(weight)
    if not total_weight:
        return None
    return round(weighted_sum / total_weight, ndigits)


def maximum(values):
    vals = [float(v) for v in values if isinstance(v, (int, float))]
    if not vals:
        return None
    value = max(vals)
    return int(value) if value.is_integer() else value


def first_number(row, *fields):
    for field in fields:
        value = row.get(field)
        if isinstance(value, (int, float)):
            return value
    return None


def first_nonblank(row, *fields):
    for field in fields:
        value = row.get(field)
        if value not in ("", None):
            return value
    return None


def parse_garmin_activity_id(path: Path):
    numbers = re.findall(r"\d{10,}", path.stem)
    return int(numbers[0]) if numbers else ""


def fit_sha256(path: Path):
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def first_session(messages):
    sessions = messages.get("session_mesgs", []) if messages else []
    return sessions[0] if sessions else {}


def path_date_code(path: Path):
    for match in re.finditer(r"(\d{4})(\d{2})(\d{2})", path.stem):
        month = int(match.group(2))
        day = int(match.group(3))
        if 1 <= month <= 12 and 1 <= day <= 31:
            return "".join(match.groups())
    return ""


def fit_activity_date_code(path: Path, messages=None):
    if messages is None:
        existing = path_date_code(path)
        if existing:
            return existing
        messages = decode_fit(path)
    session = first_session(messages)
    start = fit_datetime(session.get("start_time") or session.get("timestamp"))
    if start:
        return start.astimezone().strftime("%Y%m%d")
    return path_date_code(path) or dt.date.today().strftime("%Y%m%d")


def fit_activity_identifier(path: Path):
    activity_id = parse_garmin_activity_id(path)
    if activity_id:
        return str(activity_id)
    return f"FIT{fit_sha256(path)[:12]}"


def output_file_stem(path: Path, messages=None):
    if messages is None and path_date_code(path) and parse_garmin_activity_id(path):
        return f"{path_date_code(path)}_{fit_activity_identifier(path)}"
    return f"{fit_activity_date_code(path, messages)}_{fit_activity_identifier(path)}"


def output_month_label(path: Path, messages=None):
    if re.fullmatch(r"\d{4}-\d{2}", path.parent.name):
        return path.parent.name
    existing = path_date_code(path)
    if existing:
        return f"{existing[:4]}-{existing[4:6]}"
    date_code = fit_activity_date_code(path, messages)
    return f"{date_code[:4]}-{date_code[4:6]}"


def default_output_path(fit_path: Path):
    messages = None if path_date_code(fit_path) and parse_garmin_activity_id(fit_path) else decode_fit(fit_path)
    return DEFAULT_OUTPUT_DIR / output_month_label(fit_path, messages) / f"{WORKBOOK_VERSION_NAME}_{output_file_stem(fit_path, messages)}.xlsx"


def load_dropdown_options(path=DROPDOWN_CONFIG_PATH):
    options = {key: list(value) for key, value in DEFAULT_DROPDOWN_OPTIONS.items()}
    options["workout_focus_map"] = {}
    sqlite_options = load_dropdown_options_from_sqlite(SQLITE_DB_PATH)
    if sqlite_options:
        return sqlite_options
    if not path.exists():
        return options
    try:
        loaded = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as error:
        print(f"Dropdown config ignored: {error}")
        return options

    for key, default_values in DEFAULT_DROPDOWN_OPTIONS.items():
        if key == "shoes":
            continue
        values = loaded.get(key)
        if isinstance(values, list):
            cleaned = [str(value).strip() for value in values if str(value).strip()]
            if cleaned:
                options[key] = cleaned
        if not options.get(key):
            options[key] = list(default_values)
    options["shoes"] = []
    if isinstance(loaded.get("workout_focus_map"), dict):
        options["workout_focus_map"] = {
            str(key).strip(): [str(item).strip() for item in value if str(item).strip()]
            for key, value in loaded["workout_focus_map"].items()
            if str(key).strip() and isinstance(value, list)
        }
    return options


def load_dropdown_options_from_sqlite(db_path):
    if not db_path.exists():
        return None
    try:
        connection = sqlite3.connect(db_path)
        connection.row_factory = sqlite3.Row
    except sqlite3.DatabaseError:
        return None
    try:
        options = {key: list(value) for key, value in DEFAULT_DROPDOWN_OPTIONS.items()}
        options["workout_focus_map"] = {}

        shoe_rows = connection.execute(
            """
            SELECT shoe_code, model, nickname, is_active
            FROM shoe
            ORDER BY is_active DESC, model, nickname, shoe_code
            """
        ).fetchall()
        shoes = []
        seen_shoes = set()
        for row in shoe_rows:
            model = str(row["model"] or "").strip()
            nickname = str(row["nickname"] or "").strip()
            if model and nickname:
                label = f"{model} {nickname}".strip()
            elif model:
                label = model
            elif nickname:
                label = nickname
            else:
                label = str(row["shoe_code"] or "").strip()
            key = label.lower()
            if not label or key in seen_shoes:
                continue
            seen_shoes.add(key)
            shoes.append(label)
        if shoes:
            options["shoes"] = shoes

        workout_rows = connection.execute(
            """
            SELECT name_en, name_zh
            FROM workout_type
            ORDER BY COALESCE(sort_order, 9999), name_en
            """
        ).fetchall()
        workout_types = [str(row["name_en"]).strip() for row in workout_rows if str(row["name_en"] or "").strip()]
        options["workout_type_labels"] = {
            str(row["name_en"]).strip(): str(row["name_zh"] or row["name_en"]).strip()
            for row in workout_rows
            if str(row["name_en"] or "").strip()
        }
        if workout_types:
            options["workout_types"] = workout_types

        purpose_rows = connection.execute(
            """
            SELECT name_en, name_zh
            FROM training_purpose
            ORDER BY COALESCE(sort_order, 9999), name_en
            """
        ).fetchall()
        training_focus = [str(row["name_en"]).strip() for row in purpose_rows if str(row["name_en"] or "").strip()]
        options["training_focus_labels"] = {
            str(row["name_en"]).strip(): str(row["name_zh"] or row["name_en"]).strip()
            for row in purpose_rows
            if str(row["name_en"] or "").strip()
        }
        if training_focus:
            options["training_focus"] = training_focus

        feedback_rows = connection.execute(
            """
            SELECT dictionary_key, label
            FROM feedback_dictionary_option
            ORDER BY dictionary_key, id
            """
        ).fetchall()
        feedback_groups = {}
        for row in feedback_rows:
            key = str(row["dictionary_key"] or "").strip()
            label = str(row["label"] or "").strip()
            if not key or not label:
                continue
            feedback_groups.setdefault(key, []).append(label)
        for key in ("garmin_rpe", "garmin_feel"):
            if feedback_groups.get(key):
                options[key] = feedback_groups[key]

        map_rows = connection.execute(
            """
            SELECT
                wt.name_en AS workout_name,
                primary_purpose.name_en AS primary_name,
                secondary_purpose.name_en AS secondary_name
            FROM workout_type wt
            LEFT JOIN workout_type_training_purpose_map map
                ON map.workout_type_id = wt.id
            LEFT JOIN training_purpose primary_purpose
                ON primary_purpose.id = map.primary_training_purpose_id
            LEFT JOIN training_purpose secondary_purpose
                ON secondary_purpose.id = map.secondary_training_purpose_id
            ORDER BY COALESCE(wt.sort_order, 9999), wt.name_en
            """
        ).fetchall()
        workout_focus_map = {}
        for row in map_rows:
            workout_name = str(row["workout_name"] or "").strip()
            if not workout_name:
                continue
            purposes = []
            for purpose_name in (row["primary_name"], row["secondary_name"]):
                label = str(purpose_name or "").strip()
                if label and label not in purposes:
                    purposes.append(label)
            workout_focus_map[workout_name] = purposes
        if workout_focus_map:
            options["workout_focus_map"] = workout_focus_map

        return options
    except sqlite3.DatabaseError:
        return None
    finally:
        try:
            connection.close()
        except Exception:
            pass


def load_feedback_dictionary_seed():
    defaults = {
        "garmin_rpe": list(DEFAULT_DROPDOWN_OPTIONS["garmin_rpe"]),
        "garmin_feel": list(DEFAULT_DROPDOWN_OPTIONS["garmin_feel"]),
    }
    if not FEEDBACK_DICTIONARY_SEED_PATH.exists():
        return defaults
    try:
        loaded = json.loads(FEEDBACK_DICTIONARY_SEED_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return defaults
    for key in ("garmin_rpe", "garmin_feel"):
        values = loaded.get(key)
        if isinstance(values, list):
            cleaned = [str(value).strip() for value in values if str(value).strip()]
            if cleaned:
                defaults[key] = cleaned
    return defaults


def label_primary(value):
    value = str(value or "").strip()
    if not value:
        return ""
    return re.split(r"[（(]", value, maxsplit=1)[0].strip()


def canonical_label(value):
    return label_primary(value).replace("₂", "2").replace("₂", "2")


KNOWN_SHOE_BRANDS = (
    "adidas",
    "asics",
    "brooks",
    "hoka",
    "mizuno",
    "new balance",
    "nike",
    "on",
    "puma",
    "saucony",
)


def parse_shoe_brand_model(label):
    primary = label_primary(label) or str(label or "").strip()
    normalized = re.sub(r"\s+", " ", primary).strip()
    lowered = normalized.lower()
    for brand in KNOWN_SHOE_BRANDS:
        prefix = f"{brand} "
        if lowered.startswith(prefix):
            model = normalized[len(prefix):].strip()
            if model:
                return normalized[: len(brand)].upper() if brand == "hoka" else normalized[: len(brand)], model
    return "", normalized


def code_from_label(value, prefix):
    label = canonical_label(value).lower()
    code = re.sub(r"[^a-z0-9]+", "_", label).strip("_")
    if code:
        return code
    digest = hashlib.sha1(str(value or "").encode("utf-8")).hexdigest()[:8]
    return f"{prefix}_{digest}"


def exact_or_primary_lookup(mapping, label):
    if not label:
        return None
    if label in mapping:
        return mapping[label]
    primary = canonical_label(label)
    return mapping.get(primary)


def shoe_dimension_row(label):
    row = exact_or_primary_lookup(SHOE_DIMENSION_DEFAULTS, label)
    if row:
        return {
            **row,
            "is_active": 1,
        }
    brand, model = parse_shoe_brand_model(label)
    primary = model or label_primary(label) or str(label or "").strip()
    return {
        "shoe_code": code_from_label(primary, "shoe"),
        "brand": brand,
        "model": primary or "",
        "nickname": None,
        "category": "",
        "is_active": 1,
    }


def workout_type_dimension_row(label):
    row = exact_or_primary_lookup(WORKOUT_TYPE_DIMENSION_DEFAULTS, label)
    if row:
        code, name_en, name_zh, intensity, quality, long_run, recovery, sort_order, color = row
    else:
        name_en = label_primary(label) or str(label or "").strip() or "Other"
        code, name_zh, intensity, quality, long_run, recovery, sort_order, color = (
            code_from_label(name_en, "workout"),
            name_en,
            "Moderate",
            0,
            0,
            0,
            999,
            "#999999",
        )
    return {
        "workout_type_code": code,
        "name_en": name_en,
        "name_zh": name_zh,
        "description": None,
        "intensity_category": intensity,
        "is_quality_session": quality,
        "is_long_run": long_run,
        "is_recovery_focused": recovery,
        "sort_order": sort_order,
        "display_color": color,
    }


def training_purpose_dimension_row(label):
    row = exact_or_primary_lookup(TRAINING_PURPOSE_DIMENSION_DEFAULTS, label)
    if row:
        code, name_en, name_zh, category, physiological, recovery, performance, sort_order, color = row
    else:
        name_en = label_primary(label) or str(label or "").strip() or "Maintenance"
        code, name_zh, category, physiological, recovery, performance, sort_order, color = (
            code_from_label(name_en, "purpose"),
            name_en,
            "Maintenance",
            0,
            0,
            0,
            999,
            "#999999",
        )
    return {
        "training_purpose_code": code,
        "name_en": name_en,
        "name_zh": name_zh,
        "description": None,
        "purpose_category": category,
        "is_primary_physiological": physiological,
        "is_recovery_related": recovery,
        "is_performance_related": performance,
        "sort_order": sort_order,
        "display_color": color,
    }


def semicircles_to_degrees(value):
    if not isinstance(value, (int, float)):
        return None
    return float(value) * 180.0 / 2**31


def decode_fit(path: Path):
    stream = Stream.from_file(str(path))
    messages, errors = Decoder(stream).read()
    if errors:
        raise RuntimeError(f"FIT decode errors: {errors}")
    return messages


def compass_direction(degrees):
    if not isinstance(degrees, (int, float)):
        return ""
    labels = [
        "北風",
        "東北偏北風",
        "東北風",
        "東北偏東風",
        "東風",
        "東南偏東風",
        "東南風",
        "東南偏南風",
        "南風",
        "西南偏南風",
        "西南風",
        "西南偏西風",
        "西風",
        "西北偏西風",
        "西北風",
        "西北偏北風",
    ]
    index = int((float(degrees) + 11.25) // 22.5) % 16
    return labels[index]


def weather_code_description(code):
    descriptions = {
        0: "晴",
        1: "大致晴朗",
        2: "局部多雲",
        3: "陰",
        45: "霧",
        48: "霧淞",
        51: "毛毛雨",
        53: "毛毛雨",
        55: "毛毛雨",
        56: "凍毛毛雨",
        57: "凍毛毛雨",
        61: "小雨",
        63: "中雨",
        65: "大雨",
        66: "凍雨",
        67: "凍雨",
        71: "小雪",
        73: "中雪",
        75: "大雪",
        77: "雪粒",
        80: "陣雨",
        81: "陣雨",
        82: "強陣雨",
        85: "陣雪",
        86: "強陣雪",
        95: "雷雨",
        96: "雷雨伴冰雹",
        99: "雷雨伴冰雹",
    }
    if not isinstance(code, (int, float)):
        return ""
    return descriptions.get(int(code), "")


def nearest_hour_index(times, target):
    best_index = None
    best_seconds = None
    for index, value in enumerate(times):
        try:
            current = dt.datetime.fromisoformat(value.replace("Z", "+00:00"))
        except ValueError:
            continue
        if current.tzinfo is None:
            current = current.replace(tzinfo=dt.timezone.utc)
        delta = abs((current.astimezone(dt.timezone.utc) - target).total_seconds())
        if best_seconds is None or delta < best_seconds:
            best_index = index
            best_seconds = delta
    return best_index


def activity_location(session, records):
    gps = activity_gps_points(session, records)
    lat = gps["start_latitude"] if gps["start_latitude"] is not None else gps["end_latitude"]
    lon = gps["start_longitude"] if gps["start_longitude"] is not None else gps["end_longitude"]
    if lat is not None and lon is not None:
        return lat, lon
    return None, None


def activity_gps_points(session, records):
    def record_coordinates(record):
        return (
            semicircles_to_degrees(record.get("position_lat")),
            semicircles_to_degrees(record.get("position_long")),
        )

    start_lat = semicircles_to_degrees(session.get("start_position_lat"))
    start_lon = semicircles_to_degrees(session.get("start_position_long"))
    end_lat = semicircles_to_degrees(session.get("end_position_lat"))
    end_lon = semicircles_to_degrees(session.get("end_position_long"))

    if start_lat is None or start_lon is None:
        for record in records:
            if not isinstance(record, dict):
                continue
            start_lat, start_lon = record_coordinates(record)
            if start_lat is not None and start_lon is not None:
                break

    if end_lat is None or end_lon is None:
        for record in reversed(records):
            if not isinstance(record, dict):
                continue
            end_lat, end_lon = record_coordinates(record)
            if end_lat is not None and end_lon is not None:
                break

    return {
        "start_latitude": start_lat,
        "start_longitude": start_lon,
        "end_latitude": end_lat,
        "end_longitude": end_lon,
    }


def fetch_weather_for_activity(session, records):
    start = fit_datetime(session.get("start_time") or session.get("timestamp"))
    if start is None:
        return {}
    start_utc = start.astimezone(dt.timezone.utc)
    gps = activity_gps_points(session, records)
    latitude = gps["start_latitude"] if gps["start_latitude"] is not None else gps["end_latitude"]
    longitude = gps["start_longitude"] if gps["start_longitude"] is not None else gps["end_longitude"]
    if latitude is None or longitude is None:
        return {}

    query = {
        "latitude": round(latitude, 6),
        "longitude": round(longitude, 6),
        "start_date": start_utc.date().isoformat(),
        "end_date": start_utc.date().isoformat(),
        "hourly": ",".join(
            [
                "temperature_2m",
                "relative_humidity_2m",
                "wind_speed_10m",
                "wind_direction_10m",
                "weather_code",
            ]
        ),
        "timezone": "UTC",
        "wind_speed_unit": "kmh",
    }
    url = "https://archive-api.open-meteo.com/v1/archive?" + urlencode(query)
    with urlopen(url, timeout=15) as response:
        payload = json.load(response)

    hourly = payload.get("hourly") or {}
    index = nearest_hour_index(hourly.get("time") or [], start_utc)
    if index is None:
        return {}

    def hourly_value(name):
        values = hourly.get(name) or []
        if index >= len(values):
            return None
        return values[index]

    wind_degrees = hourly_value("wind_direction_10m")
    wind_label = compass_direction(wind_degrees)
    weather = {
        "weather_temp": rounded(hourly_value("temperature_2m"), 1),
        "humidity": rounded(hourly_value("relative_humidity_2m"), 0),
        "wind_direction": f"{round(wind_degrees)}° ({wind_label})" if wind_degrees is not None else "",
        "wind_speed": f"{rounded(hourly_value('wind_speed_10m'), 1)} km/h"
        if hourly_value("wind_speed_10m") is not None
        else "",
        "weather_description": weather_code_description(hourly_value("weather_code")),
    }
    return {key: value for key, value in weather.items() if value not in ("", None)}


def records_for_lap(records, start_time, elapsed_seconds):
    if not start_time or not elapsed_seconds:
        return []
    end_time = start_time + dt.timedelta(seconds=float(elapsed_seconds))
    return [
        record
        for record in records
        if (timestamp := fit_datetime(record.get("timestamp")))
        and start_time <= timestamp <= end_time
    ]


def stamina_at(records, fallback=None):
    for record in records:
        value = first_number(record, *STAMINA_RECORD_FIELDS)
        if value is not None:
            return int(value)
    return fallback


def has_stamina_data(messages):
    session = first_session(messages)
    if first_number(session, STAMINA_SESSION_START, *STAMINA_SESSION_END_FIELDS) is not None:
        return True
    for record in messages.get("record_mesgs", []):
        if first_number(record, *STAMINA_RECORD_FIELDS) is not None:
            return True
    return False


def first_message_number(messages, message_names, *fields):
    for message_name in message_names:
        for message in messages.get(message_name, []):
            value = first_number(message, *fields)
            if value is not None:
                return value
    return None


def garmin_rpe_label(value, rpe_options):
    if not isinstance(value, (int, float)):
        return ""
    rating = int(round(float(value) / 10))
    if 1 <= rating <= len(rpe_options):
        return rpe_options[rating - 1]
    return str(value)


def normalize_rpe(value, rpe_options):
    if value in ("", None):
        return ""
    if isinstance(value, str):
        if value in rpe_options:
            return value
        for option in rpe_options:
            label = option.split(" - ", 1)[-1]
            if value == label or value == label.split(" (", 1)[0]:
                return option
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return str(value)

    rating = int(round(numeric / 10)) if numeric > len(rpe_options) else int(round(numeric))
    if 1 <= rating <= len(rpe_options):
        return rpe_options[rating - 1]
    return str(value)


def garmin_feel_label(value, feel_options):
    if not isinstance(value, (int, float)):
        return ""
    scores = [0, 25, 50, 75, 100]
    index = min(range(len(scores)), key=lambda idx: abs(float(value) - scores[idx]))
    if index < len(feel_options):
        return feel_options[index]
    return str(value)


def normalize_feel(value, feel_options):
    if value in ("", None):
        return ""
    if isinstance(value, str):
        if value in feel_options:
            return value
        for option in feel_options:
            if value == option.split(" (", 1)[0]:
                return option
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return str(value)
    return garmin_feel_label(numeric, feel_options)


def build_rows(messages):
    laps = messages.get("lap_mesgs", [])
    records = messages.get("record_mesgs", [])
    sessions = messages.get("session_mesgs", [])
    session = sessions[0] if sessions else {}

    rows = []
    for index, lap in enumerate(laps, start=1):
        distance = float(lap.get("total_distance") or 0)
        elapsed = float(lap.get("total_timer_time") or lap.get("total_elapsed_time") or 0)
        lap_start = fit_datetime(lap.get("start_time"))
        lap_records = records_for_lap(records, lap_start, elapsed)

        cadence = first_number(lap, "avg_running_cadence", "avg_cadence")
        fractional_cadence = first_number(lap, "avg_fractional_cadence") or 0
        cadence_spm = (float(cadence) + float(fractional_cadence)) * 2 if cadence is not None else None

        start_stamina = stamina_at(lap_records)
        end_stamina = stamina_at(reversed(lap_records))
        if index == 1:
            start_stamina = int(first_number(session, STAMINA_SESSION_START) or start_stamina or 0)
        if index == len(laps):
            end_stamina = int(first_number(session, *STAMINA_SESSION_END_FIELDS) or end_stamina or 0)

        avg_heart_rate = first_number(lap, "avg_heart_rate")
        if avg_heart_rate is None:
            avg_heart_rate = average([record.get("heart_rate") for record in lap_records], 1)
        max_heart_rate = first_number(lap, "max_heart_rate")
        if max_heart_rate is None:
            max_heart_rate = maximum([record.get("heart_rate") for record in lap_records])
        avg_power = first_number(lap, "avg_power")
        if avg_power is None:
            avg_power = average([record.get("power") for record in lap_records], 1)

        rows.append(
            [
                index,
                round(distance),
                round(elapsed),
                pace_text(elapsed, distance),
                rounded(avg_heart_rate, 1),
                None,
                max_heart_rate,
                rounded(cadence_spm, 1),
                rounded(avg_power, 1),
                None,
                rounded(lap.get("avg_vertical_oscillation"), 1),
                rounded(lap.get("avg_vertical_ratio"), 1),
                rounded(lap.get("avg_stance_time"), 1),
                rounded(lap.get("avg_step_length"), 1),
                rounded(average([r.get("temperature") for r in lap_records], 1), 1)
                if lap_records
                else rounded(lap.get("avg_temperature"), 1),
                start_stamina,
                end_stamina,
                rounded(lap.get("total_ascent") or 0, 1),
            ]
        )
    return rows, session


def prompt_choice(label, options):
    print(f"\n{label}")
    for index, option in enumerate(options, start=1):
        print(f"  {index}. {option}")
    raw = input("請輸入編號或直接輸入文字，留空略過: ").strip()
    if not raw:
        return ""
    if raw.isdigit() and 1 <= int(raw) <= len(options):
        return options[int(raw) - 1]
    return raw


def prompt_text(label):
    return input(f"{label}，留空略過: ").strip()


def collect_metadata(args, dropdown_options):
    metadata = {
        "activity_name": args.activity_name or "",
        "shoe": args.shoe or "",
        "weather_temp": args.weather_temp if args.weather_temp is not None else "",
        "humidity": args.humidity if args.humidity is not None else "",
        "wind_direction": args.wind_direction or "",
        "wind_speed": args.wind_speed or "",
        "weather_description": args.weather_description or "",
        "workout_type": args.workout_type or "",
        "training_focus": args.training_focus or "",
        "feel": normalize_feel(args.feel, dropdown_options["garmin_feel"]),
        "rpe": normalize_rpe(args.rpe, dropdown_options["garmin_rpe"]),
        "fueling": args.fueling or "",
        "max_hr": args.max_hr if args.max_hr is not None else "",
        "critical_power": args.critical_power if args.critical_power is not None else "",
        "training_effect_aerobic": args.training_effect_aerobic if args.training_effect_aerobic is not None else "",
        "training_effect_anaerobic": args.training_effect_anaerobic if args.training_effect_anaerobic is not None else "",
        "training_load": args.training_load if args.training_load is not None else "",
        "recovery_time_hr": args.recovery_time_hr if args.recovery_time_hr is not None else "",
        "notes": args.notes or "",
    }
    if not args.interactive:
        return metadata

    if not metadata["activity_name"]:
        metadata["activity_name"] = prompt_text("活動名稱")
    if not metadata["shoe"]:
        metadata["shoe"] = prompt_choice("鞋款", dropdown_options["shoes"])
    if not args.fetch_weather and metadata["weather_temp"] == "":
        metadata["weather_temp"] = prompt_text("天氣氣溫(°C)")
    if not args.fetch_weather and metadata["humidity"] == "":
        metadata["humidity"] = prompt_text("濕度(%)")
    if not args.fetch_weather and not metadata["wind_direction"]:
        metadata["wind_direction"] = prompt_text("風向")
    if not args.fetch_weather and not metadata["wind_speed"]:
        metadata["wind_speed"] = prompt_text("風速")
    if not args.fetch_weather and not metadata["weather_description"]:
        metadata["weather_description"] = prompt_text("天氣描述")
    if not metadata["workout_type"]:
        metadata["workout_type"] = prompt_choice("課表類型", dropdown_options["workout_types"])
    if not metadata["training_focus"]:
        metadata["training_focus"] = prompt_choice("訓練目的（Training Focus）", dropdown_options["training_focus"])
    if metadata["feel"] == "":
        metadata["feel"] = prompt_text("感覺如何")
    if metadata["rpe"] == "":
        metadata["rpe"] = prompt_text("感受難度(1-10)")
    if not metadata["fueling"]:
        metadata["fueling"] = prompt_text("補給紀錄")
    if metadata["max_hr"] == "":
        metadata["max_hr"] = prompt_text("最大心率")
    if metadata["critical_power"] == "":
        metadata["critical_power"] = prompt_text("Critical Power(W)")
    if metadata["training_effect_aerobic"] == "":
        metadata["training_effect_aerobic"] = prompt_text("Training Effect (Aerobic)")
    if metadata["training_effect_anaerobic"] == "":
        metadata["training_effect_anaerobic"] = prompt_text("Training Effect (Anaerobic)")
    if metadata["training_load"] == "":
        metadata["training_load"] = prompt_text("Training Load")
    if metadata["recovery_time_hr"] == "":
        metadata["recovery_time_hr"] = prompt_text("Recovery Time (hr)")
    if not metadata["notes"]:
        metadata["notes"] = prompt_text("備註")
    return metadata


def apply_auto_weather(metadata, session, records, enabled):
    if not enabled:
        return metadata
    if all(metadata.get(key) not in ("", None) for key in WEATHER_FIELDS):
        return metadata
    try:
        weather = fetch_weather_for_activity(session, records)
    except Exception as error:
        print(f"Weather lookup skipped: {error}")
        return metadata
    result = dict(metadata)
    for key, value in weather.items():
        if result.get(key) in ("", None):
            result[key] = value
    return result


def apply_fit_metadata(metadata, messages):
    result = dict(metadata)
    if result.get("max_hr") in ("", None):
        value = first_message_number(
            messages,
            ("zones_target_mesgs", "time_in_zone_mesgs"),
            "max_heart_rate",
        )
        if value is not None:
            result["max_hr"] = value
    if result.get("critical_power") in ("", None):
        value = first_message_number(
            messages,
            ("zones_target_mesgs", "time_in_zone_mesgs"),
            "functional_threshold_power",
        )
        if value is not None:
            result["critical_power"] = value
    return result


def apply_file_identity(metadata, fit_path: Path):
    result = dict(metadata)
    result["garmin_activity_id"] = parse_garmin_activity_id(fit_path)
    result["fit_sha256"] = fit_sha256(fit_path)
    return result


def coerce_metadata(metadata):
    result = dict(metadata)
    for key in (
        "weather_temp",
        "humidity",
        "rpe",
        "max_hr",
        "critical_power",
        "training_effect_aerobic",
        "training_effect_anaerobic",
        "training_load",
        "recovery_time_hr",
    ):
        if result.get(key) == "":
            continue
        try:
            value = float(result.get(key, ""))
            result[key] = int(value) if value.is_integer() else value
        except (TypeError, ValueError):
            pass
    return result


def finalized_metadata(metadata, messages, session, rows, fit_path, fetch_weather, dropdown_options):
    result = apply_file_identity(metadata or {}, fit_path)
    result = apply_fit_metadata(result, messages)
    result = apply_auto_weather(result, session, messages.get("record_mesgs", []), fetch_weather)
    result = coerce_metadata(result)

    if result.get("rpe", "") == "":
        result["rpe"] = garmin_rpe_label(session.get("workout_rpe"), dropdown_options["garmin_rpe"])
    else:
        result["rpe"] = normalize_rpe(result.get("rpe"), dropdown_options["garmin_rpe"])
    if result.get("feel", "") == "":
        result["feel"] = garmin_feel_label(session.get("workout_feel"), dropdown_options["garmin_feel"])
    else:
        result["feel"] = normalize_feel(result.get("feel"), dropdown_options["garmin_feel"])
    if result.get("training_effect_aerobic", "") == "":
        value = session.get("total_training_effect")
        if isinstance(value, (int, float)):
            result["training_effect_aerobic"] = round(float(value), 1)
    if result.get("training_effect_anaerobic", "") == "":
        value = session.get("total_anaerobic_training_effect")
        if isinstance(value, (int, float)):
            result["training_effect_anaerobic"] = round(float(value), 1)
    if result.get("training_load", "") == "":
        value = session.get("training_load_peak")
        if isinstance(value, (int, float)):
            result["training_load"] = round(float(value))
    return coerce_metadata(result)


def activity_date(session, fit_path: Path):
    start = fit_datetime(session.get("start_time") or session.get("timestamp"))
    if start:
        local = start.astimezone()
        return local.strftime("%Y/%m/%d")
    return fit_path.stem


def activity_type(session):
    profile_name = session.get("sport_profile_name")
    if profile_name not in ("", None):
        return str(profile_name)

    sport = session.get("sport")
    sub_sport = session.get("sub_sport")
    sport_labels = {
        "running": "跑步",
        "cycling": "自行車",
        "walking": "健走",
        "hiking": "健行",
        "swimming": "游泳",
    }
    sub_sport_labels = {
        "trail": "越野跑",
        "trail_running": "越野跑",
        "treadmill": "跑步機",
        "track": "田徑場跑步",
    }
    if sub_sport in sub_sport_labels:
        return sub_sport_labels[sub_sport]
    if sport in sport_labels:
        return sport_labels[sport]
    values = [str(value) for value in (sport, sub_sport) if value not in ("", None, "generic")]
    return " / ".join(values)


def activity_name(session, metadata):
    if metadata.get("activity_name") not in ("", None):
        return metadata.get("activity_name")
    for key in ("activity_name", "name", "workout_name", "title"):
        value = session.get(key)
        if value not in ("", None):
            return value
    return ""


def duration_text(seconds):
    if not isinstance(seconds, (int, float)):
        return ""
    total = int(round(float(seconds)))
    hours, remainder = divmod(total, 3600)
    minutes, seconds = divmod(remainder, 60)
    if hours:
        return f"{hours}:{minutes:02d}:{seconds:02d}"
    return f"{minutes}:{seconds:02d}"


def activity_start_iso(session):
    start = fit_datetime(session.get("start_time") or session.get("timestamp"))
    if start is None:
        return None
    return start.astimezone().strftime("%Y-%m-%dT%H:%M:%S")


def wind_direction_degrees(value):
    if value in ("", None):
        return None
    if isinstance(value, (int, float)):
        return round(float(value), 1)
    match = re.search(r"-?\d+(?:\.\d+)?", str(value))
    return round(float(match.group(0)), 1) if match else None


def wind_speed_mps(value):
    if value in ("", None):
        return None
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    text = str(value).strip().lower()
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    speed = float(match.group(0))
    if "km" in text:
        speed = speed / 3.6
    return round(speed, 2)


def activity_summary(rows):
    total_distance = sum(row[1] for row in rows if isinstance(row[1], (int, float)))
    total_seconds = sum(row[2] for row in rows if isinstance(row[2], (int, float)))
    return {
        "distance_km": round(total_distance / 1000, 2) if total_distance else "",
        "duration": duration_text(total_seconds),
        "avg_pace": pace_text(total_seconds, total_distance) if total_distance and total_seconds else "",
    }


def session_running_cadence_spm(session):
    cadence = first_number(session, "avg_running_cadence", "avg_cadence")
    if cadence is None:
        return None
    fractional_cadence = first_number(session, "avg_fractional_cadence") or 0
    return rounded((float(cadence) + float(fractional_cadence)) * 2, 1)


def metric_summary_value(session, session_field, rows, row_index, ndigits=1):
    value = first_number(session, session_field)
    if value is not None:
        return rounded(value, ndigits)
    return weighted_average(((row[row_index], row[2]) for row in rows), ndigits)


def running_economy_summary(rows, session=None):
    session = session or {}
    return {
        "avg_cadence": session_running_cadence_spm(session) or weighted_average(((row[7], row[2]) for row in rows), 1),
        "avg_step_length": metric_summary_value(session, "avg_step_length", rows, 13, 1),
        "avg_gct": metric_summary_value(session, "avg_stance_time", rows, 12, 1),
        "avg_vertical_oscillation": metric_summary_value(session, "avg_vertical_oscillation", rows, 10, 1),
        "avg_vertical_ratio": metric_summary_value(session, "avg_vertical_ratio", rows, 11, 1),
    }


def stamina_summary(rows):
    if not rows:
        return "", ""
    return rows[0][15], rows[-1][16]


def apply_styles(ws, last_row):
    blue = PatternFill("solid", fgColor="1F4E78")
    total_fill = PatternFill("solid", fgColor="D9EAF7")
    thin_gray = Side(style="thin", color="D9E2F3")

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    title = ws["A1"]
    title.fill = blue
    title.font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    for cell in ws[2]:
        cell.fill = blue
        cell.font = Font(name="Noto Sans CJK SC", size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(top=thin_gray, bottom=thin_gray)

    for row in ws.iter_rows(min_row=3, max_row=last_row, max_col=len(HEADERS)):
        for cell in row:
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(bottom=thin_gray)

    for cell in ws[last_row]:
        cell.fill = total_fill
        cell.font = Font(name="Arial", size=10, bold=True)

    widths = {
        "A": 7,
        "B": 9,
        "C": 9,
        "D": 13,
        "E": 9,
        "F": 10,
        "G": 9,
        "H": 12,
        "I": 10,
        "J": 9,
        "K": 11,
        "L": 9,
        "M": 11,
        "N": 10,
        "O": 9,
        "P": 10,
        "Q": 10,
        "R": 9,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for col in range(2, len(HEADERS) + 1):
        for row in range(3, last_row + 1):
            ws.cell(row, col).number_format = "0.0"
    for col in (1, 2, 3, 7, 16, 17):
        for row in range(3, last_row + 1):
            ws.cell(row, col).number_format = "0"
    for col in (6, 10):
        for row in range(3, last_row + 1):
            ws.cell(row, col).number_format = "0.0%"

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(HEADERS))}{last_row}"


def add_options_sheet(wb, dropdown_options):
    ws = wb.create_sheet("選項")
    columns = [
        ("鞋款", dropdown_options["shoes"]),
        ("課表類型", dropdown_options["workout_types"]),
        ("訓練目的", dropdown_options["training_focus"]),
        ("感受難度", dropdown_options["garmin_rpe"]),
        ("感覺如何", dropdown_options["garmin_feel"]),
    ]
    for col, (title, options) in enumerate(columns, start=1):
        ws.cell(1, col, title)
        ws.cell(1, col).font = Font(name="Arial", bold=True)
        for row, option in enumerate(options, start=2):
            ws.cell(row, col, option)
        ws.column_dimensions[get_column_letter(col)].width = max(12, max(len(x) for x in [title, *options]) + 2)
    ws.sheet_state = "hidden"
    return ws


def add_metadata_sheet(wb, metadata, fit_path, session, rows, records, dropdown_options):
    metadata = coerce_metadata(metadata)
    if metadata.get("rpe", "") == "":
        metadata["rpe"] = garmin_rpe_label(session.get("workout_rpe"), dropdown_options["garmin_rpe"])
    else:
        metadata["rpe"] = normalize_rpe(metadata.get("rpe"), dropdown_options["garmin_rpe"])
    if metadata.get("feel", "") == "":
        metadata["feel"] = garmin_feel_label(session.get("workout_feel"), dropdown_options["garmin_feel"])
    else:
        metadata["feel"] = normalize_feel(metadata.get("feel"), dropdown_options["garmin_feel"])
    if metadata.get("training_effect_aerobic", "") == "":
        value = session.get("total_training_effect")
        if isinstance(value, (int, float)):
            metadata["training_effect_aerobic"] = round(float(value), 1)
    if metadata.get("training_effect_anaerobic", "") == "":
        value = session.get("total_anaerobic_training_effect")
        if isinstance(value, (int, float)):
            metadata["training_effect_anaerobic"] = round(float(value), 1)
    if metadata.get("training_load", "") == "":
        value = session.get("training_load_peak")
        if isinstance(value, (int, float)):
            metadata["training_load"] = round(float(value))

    start = fit_datetime(session.get("start_time") or session.get("timestamp"))
    gps = activity_gps_points(session, records)
    activity = activity_summary(rows)
    economy = running_economy_summary(rows, session)
    stamina_start, stamina_end = stamina_summary(rows)
    metadata_sections = [
        (
            "Metadata",
            "1F4E78",
            "D9EAF7",
            [
                ("Excel Schema Version", "v1.1", "excel_schema_version"),
                ("資料來源", fit_path.name, "source"),
                ("Garmin Activity ID", metadata.get("garmin_activity_id", ""), "garmin_activity_id"),
                ("FIT Hash (SHA-256)", metadata.get("fit_sha256", ""), "fit_sha256"),
            ],
        ),
        (
            "Activity",
            "2E7D32",
            "E2F0D9",
            [
                ("活動日期", activity_date(session, fit_path), "activity_date"),
                ("開始時間", start.astimezone().strftime("%H:%M:%S") if start else "", "start_time"),
                ("活動類型", activity_type(session), "activity_type"),
                ("活動名稱", activity_name(session, metadata), "activity_name"),
                ("距離 (km)", activity["distance_km"], "distance_km"),
                ("時間", activity["duration"], "duration"),
                ("平均配速", activity["avg_pace"], "avg_pace"),
                ("課表類型", metadata.get("workout_type", ""), "workout_type"),
                ("訓練目的", metadata.get("training_focus", ""), "training_focus"),
                ("鞋款", metadata.get("shoe", ""), "shoe"),
            ],
        ),
        (
            "GPS",
            "4F81BD",
            "D9EAF7",
            [
                ("起點緯度", gps["start_latitude"], "start_latitude"),
                ("起點經度", gps["start_longitude"], "start_longitude"),
                ("終點緯度", gps["end_latitude"], "end_latitude"),
                ("終點經度", gps["end_longitude"], "end_longitude"),
            ],
        ),
        (
            "Environment",
            "B8860B",
            "FFF2CC",
            [
                ("天氣氣溫 (°C)", metadata.get("weather_temp", ""), "weather_temp"),
                ("濕度 (%)", metadata.get("humidity", ""), "humidity"),
                ("風向", metadata.get("wind_direction", ""), "wind_direction"),
                ("風速", metadata.get("wind_speed", ""), "wind_speed"),
                ("天氣描述", metadata.get("weather_description", ""), "weather_description"),
            ],
        ),
        (
            "Subjective",
            "C55A11",
            "FCE4D6",
            [
                ("感覺如何", metadata.get("feel", ""), "feel"),
                ("感受難度", metadata.get("rpe", ""), "rpe"),
                ("補給紀錄", metadata.get("fueling", ""), "fueling"),
                ("備註", metadata.get("notes", ""), "notes"),
            ],
        ),
        (
            "Training Metrics",
            "7030A0",
            "EADCF8",
            [
                ("最大心率", metadata.get("max_hr", ""), "max_hr"),
                ("Critical Power (W)", metadata.get("critical_power", ""), "critical_power"),
                ("Training Effect (Aerobic)", metadata.get("training_effect_aerobic", ""), "training_effect_aerobic"),
                ("Training Effect (Anaerobic)", metadata.get("training_effect_anaerobic", ""), "training_effect_anaerobic"),
                ("Training Load", metadata.get("training_load", ""), "training_load"),
                ("Recovery Time (hr)", metadata.get("recovery_time_hr", ""), "recovery_time_hr"),
                ("Stamina 起始 (%)", stamina_start, "stamina_start"),
                ("Stamina 結束 (%)", stamina_end, "stamina_end"),
            ],
        ),
        (
            "Running Economy",
            "156082",
            "DDEBF7",
            [
                ("平均步頻", economy["avg_cadence"], "avg_cadence"),
                ("平均步幅 (mm)", economy["avg_step_length"], "avg_step_length"),
                ("平均觸地時間 GCT (ms)", economy["avg_gct"], "avg_gct"),
                ("平均垂直振幅 (mm)", economy["avg_vertical_oscillation"], "avg_vertical_oscillation"),
                ("平均垂直比", economy["avg_vertical_ratio"], "avg_vertical_ratio"),
            ],
        ),
    ]

    ws = wb.create_sheet("活動資訊", 0)
    ws["A1"] = "活動資訊"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws.merge_cells("A1:B1")

    row_by_key = {}
    section_rows = {}
    row_fills = {}
    current_row = 2
    for section_name, header_color, fill_color, rows in metadata_sections:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
        ws.cell(current_row, 1, section_name)
        ws.cell(current_row, 1).fill = PatternFill("solid", fgColor=header_color)
        ws.cell(current_row, 1).font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
        ws.cell(current_row, 1).alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[current_row].height = 22
        section_rows[current_row] = header_color
        current_row += 1
        for label, value, key in rows:
            ws.cell(current_row, 1, label)
            ws.cell(current_row, 2, cell_value(value))
            row_by_key[key] = current_row
            row_fills[current_row] = fill_color
            current_row += 1

    option_ws = add_options_sheet(wb, dropdown_options)
    validations = {
        row_by_key["shoe"]: f"='選項'!$A$2:$A${len(dropdown_options['shoes']) + 1}",
        row_by_key["workout_type"]: f"='選項'!$B$2:$B${len(dropdown_options['workout_types']) + 1}",
        row_by_key["training_focus"]: f"='選項'!$C$2:$C${len(dropdown_options['training_focus']) + 1}",
        row_by_key["rpe"]: f"='選項'!$D$2:$D${len(dropdown_options['garmin_rpe']) + 1}",
        row_by_key["feel"]: f"='選項'!$E$2:$E${len(dropdown_options['garmin_feel']) + 1}",
    }
    for row, formula in validations.items():
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(ws.cell(row, 2))

    for row in (row_by_key["max_hr"], row_by_key["critical_power"]):
        positive_number = DataValidation(
            type="decimal",
            operator="greaterThan",
            formula1="0",
            allow_blank=True,
        )
        ws.add_data_validation(positive_number)
        positive_number.add(ws.cell(row, 2))

    for row in (
        row_by_key["training_effect_aerobic"],
        row_by_key["training_effect_anaerobic"],
        row_by_key["training_load"],
        row_by_key["recovery_time_hr"],
    ):
        non_negative_number = DataValidation(
            type="decimal",
            operator="greaterThanOrEqual",
            formula1="0",
            allow_blank=True,
        )
        ws.add_data_validation(non_negative_number)
        non_negative_number.add(ws.cell(row, 2))

    thin_gray = Side(style="thin", color="D9E2F3")
    for row, fill_color in row_fills.items():
        fill = PatternFill("solid", fgColor=fill_color)
        ws.cell(row, 1).fill = fill
        ws.cell(row, 2).fill = fill
        ws.cell(row, 1).font = Font(name="Arial", bold=True)
        ws.cell(row, 1).alignment = Alignment(horizontal="left")
        ws.cell(row, 2).alignment = Alignment(horizontal="left")
        ws.cell(row, 1).border = Border(bottom=thin_gray)
        ws.cell(row, 2).border = Border(bottom=thin_gray)
    for row in section_rows:
        ws.cell(row, 1).border = Border(top=thin_gray, bottom=thin_gray)
        ws.cell(row, 2).border = Border(top=thin_gray, bottom=thin_gray)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 72
    ws.freeze_panes = "A2"
    return option_ws


def add_workout_structure_sheet(wb, messages):
    structure = sqlite_workout_structure_row(messages)
    step_rows = sqlite_workout_step_rows(messages)
    split_rows = sqlite_workout_split_rows(messages)

    ws = wb.create_sheet("課表結構")
    ws["A1"] = "課表結構"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws.merge_cells("A1:B1")

    summary_rows = [
        ("有課表結構", "Yes" if structure["has_workout_structure"] else "No"),
        ("來源", structure["source"] or ""),
        ("運動", structure["sport"] or ""),
        ("子類型", structure["sub_sport"] or ""),
        ("課表名稱", structure["workout_name"] or ""),
        ("課表描述", structure["workout_description"] or ""),
        ("有效步驟數", structure["num_valid_steps"] if structure["num_valid_steps"] is not None else ""),
        ("Workout Steps", len(step_rows)),
        ("Workout Splits", len(split_rows)),
    ]

    thin_gray = Side(style="thin", color="D9E2F3")
    summary_fill = PatternFill("solid", fgColor="D9EAF7")
    row_no = 2
    for label, value in summary_rows:
        ws.cell(row_no, 1, label)
        ws.cell(row_no, 2, cell_value(value))
        ws.cell(row_no, 1).font = Font(name="Arial", bold=True)
        ws.cell(row_no, 1).fill = summary_fill
        ws.cell(row_no, 2).fill = summary_fill
        ws.cell(row_no, 1).border = Border(bottom=thin_gray)
        ws.cell(row_no, 2).border = Border(bottom=thin_gray)
        row_no += 1

    row_no += 1
    step_header_row = row_no
    step_headers = [
        "Step",
        "Intensity",
        "Duration Type",
        "Distance (m)",
        "Time (sec)",
        "Target Type",
        "Target Low",
        "Target High",
        "Repeat Steps",
    ]
    for col, title in enumerate(step_headers, start=1):
        cell = ws.cell(step_header_row, col, title)
        cell.fill = PatternFill("solid", fgColor="2E7D32")
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for index, row in enumerate(step_rows, start=step_header_row + 1):
        values = [
            row["step_index"],
            row["intensity"],
            row["duration_type"],
            row["duration_distance_m"],
            row["duration_time_sec"],
            row["target_type"],
            row["target_value_low"] if row["target_value_low"] is not None else row["custom_target_value_low"],
            row["target_value_high"] if row["target_value_high"] is not None else row["custom_target_value_high"],
            row["repeat_steps"],
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(index, col, cell_value(value))

    row_no = step_header_row + max(len(step_rows), 1) + 3
    split_header_row = row_no
    split_headers = [
        "Split",
        "Split Type",
        "Num Splits",
        "Distance (m)",
        "Timer Time (sec)",
        "Avg Speed (m/s)",
        "Sport",
        "Sub Sport",
    ]
    for col, title in enumerate(split_headers, start=1):
        cell = ws.cell(split_header_row, col, title)
        cell.fill = PatternFill("solid", fgColor="7030A0")
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for index, row in enumerate(split_rows, start=split_header_row + 1):
        values = [
            row["split_index"],
            row["split_type"],
            row["num_splits"],
            row["total_distance_m"],
            row["total_timer_time_sec"],
            row["avg_speed_mps"],
            row["sport"],
            row["sub_sport"],
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(index, col, cell_value(value))

    widths = {
        "A": 12,
        "B": 20,
        "C": 18,
        "D": 14,
        "E": 16,
        "F": 16,
        "G": 14,
        "H": 14,
        "I": 14,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    return ws


def add_percentage_values(ws, row_count, metadata):
    max_hr = metadata.get("max_hr")
    critical_power = metadata.get("critical_power")
    for row in range(3, row_count + 3):
        if isinstance(max_hr, (int, float)) and max_hr > 0 and isinstance(ws.cell(row, 5).value, (int, float)):
            ws.cell(row, 6, ws.cell(row, 5).value / max_hr)
        if (
            isinstance(critical_power, (int, float))
            and critical_power > 0
            and isinstance(ws.cell(row, 9).value, (int, float))
        ):
            ws.cell(row, 10, ws.cell(row, 9).value / critical_power)


def add_total_row(ws, row_count, rows):
    total_row = row_count + 3
    data_start = 3
    data_end = row_count + 2
    total_distance = sum(row[1] for row in rows if isinstance(row[1], (int, float)))
    total_seconds = sum(row[2] for row in rows if isinstance(row[2], (int, float)))
    ws.cell(total_row, 1, "總計/平均")
    ws.cell(total_row, 2, f"=SUM(B{data_start}:B{data_end})")
    ws.cell(total_row, 3, f"=SUM(C{data_start}:C{data_end})")
    ws.cell(total_row, 4, pace_text(total_seconds, total_distance))
    for col in (5, 6, 8, 9, 10, 11, 12, 13, 14, 15):
        letter = get_column_letter(col)
        ws.cell(total_row, col, f"=SUMPRODUCT({letter}{data_start}:{letter}{data_end},$C{data_start}:$C{data_end})/SUM($C{data_start}:$C{data_end})")
    ws.cell(total_row, 7, f"=MAX(G{data_start}:G{data_end})")
    ws.cell(total_row, 16, f"=INDEX(P{data_start}:P{data_end},1)")
    ws.cell(total_row, 17, f"=INDEX(Q{data_start}:Q{data_end},{row_count})")
    ws.cell(total_row, 18, f"=SUM(R{data_start}:R{data_end})")
    return total_row


def add_charts(wb, row_count):
    ws = wb["每公里數據"]
    chart_ws = wb.create_sheet("圖表")
    chart_ws["A1"] = "配速 / 心率 / Stamina 趨勢圖"
    chart_ws["A1"].font = Font(name="Arial", size=14, bold=True)

    max_row = row_count + 2
    cats = Reference(ws, min_col=1, min_row=3, max_row=max_row)

    chart1 = LineChart()
    chart1.title = "心率與 Stamina 隨公里數變化"
    chart1.y_axis.title = "心率 / Stamina"
    chart1.x_axis.title = "公里"
    for col in (5, 16, 17):
        data = Reference(ws, min_col=col, min_row=2, max_row=max_row)
        chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.height = 12
    chart1.width = 24
    chart_ws.add_chart(chart1, "A3")

    chart2 = LineChart()
    chart2.title = "配速趨勢(秒/公里，數值越低代表越快)"
    chart2.y_axis.title = "秒/公里"
    chart2.x_axis.title = "公里"
    data = Reference(ws, min_col=3, min_row=2, max_row=max_row)
    chart2.add_data(data, titles_from_data=True)
    chart2.set_categories(cats)
    chart2.height = 12
    chart2.width = 24
    chart_ws.add_chart(chart2, "A28")


def sqlite_activity_row(fit_path, messages, session, rows, metadata):
    total_distance_m = first_number(session, "total_distance")
    if total_distance_m is None:
        total_distance_m = sum(row[1] for row in rows if isinstance(row[1], (int, float)))
    total_seconds = first_number(session, "total_timer_time", "total_elapsed_time")
    if total_seconds is None:
        total_seconds = sum(row[2] for row in rows if isinstance(row[2], (int, float)))

    economy = running_economy_summary(rows, session)
    stamina_start, stamina_end = stamina_summary(rows) if has_stamina_data(messages) else (None, None)
    gps = activity_gps_points(session, messages.get("record_mesgs", []))

    return {
        "fit_sha256": metadata.get("fit_sha256"),
        "garmin_activity_id": int_or_none(metadata.get("garmin_activity_id")),
        "excel_schema_version": "v1.1",
        "source_file_name": fit_path.name,
        "data_source": "FIT",
        "activity_start_time": activity_start_iso(session),
        "activity_type": activity_type(session),
        "activity_name": null_if_blank(activity_name(session, metadata)),
        "distance_km": round(float(total_distance_m) / 1000, 3) if total_distance_m else None,
        "duration_sec": int_or_none(total_seconds),
        "workout_type_id": None,
        "shoe_id": None,
        "temperature_c": float_or_none(metadata.get("weather_temp")),
        "humidity_pct": float_or_none(metadata.get("humidity")),
        "wind_speed_mps": wind_speed_mps(metadata.get("wind_speed")),
        "wind_direction_deg": wind_direction_degrees(metadata.get("wind_direction")),
        "weather_description": null_if_blank(metadata.get("weather_description")),
        "max_hr": int_or_none(metadata.get("max_hr") or first_number(session, "max_heart_rate")),
        "avg_hr": int_or_none(first_number(session, "avg_heart_rate") or weighted_average(((row[4], row[2]) for row in rows), 1)),
        "critical_power_w": int_or_none(metadata.get("critical_power")),
        "training_effect_aerobic": float_or_none(metadata.get("training_effect_aerobic")),
        "training_effect_anaerobic": float_or_none(metadata.get("training_effect_anaerobic")),
        "training_load": int_or_none(metadata.get("training_load")),
        "recovery_time_hr": float_or_none(metadata.get("recovery_time_hr")),
        "stamina_start_pct": int_or_none(stamina_start),
        "stamina_end_pct": int_or_none(stamina_end),
        "avg_cadence_spm": float_or_none(economy["avg_cadence"]),
        "avg_stride_length_mm": float_or_none(economy["avg_step_length"]),
        "avg_gct_ms": float_or_none(economy["avg_gct"]),
        "avg_vertical_oscillation_mm": float_or_none(economy["avg_vertical_oscillation"]),
        "avg_vertical_ratio_pct": float_or_none(economy["avg_vertical_ratio"]),
        "garmin_feeling": null_if_blank(metadata.get("feel")),
        "garmin_perceived_effort": null_if_blank(metadata.get("rpe")),
        "nutrition": null_if_blank(metadata.get("fueling")),
        "notes": null_if_blank(metadata.get("notes")),
        "start_latitude": gps["start_latitude"],
        "start_longitude": gps["start_longitude"],
        "end_latitude": gps["end_latitude"],
        "end_longitude": gps["end_longitude"],
    }


def sqlite_split_rows(messages):
    laps = messages.get("lap_mesgs", [])
    records = messages.get("record_mesgs", [])
    session = first_session(messages)
    stamina_supported = has_stamina_data(messages)
    result = []

    for index, lap in enumerate(laps, start=1):
        distance = float(lap.get("total_distance") or 0)
        elapsed = float(lap.get("total_timer_time") or lap.get("total_elapsed_time") or 0)
        lap_start = fit_datetime(lap.get("start_time"))
        lap_records = records_for_lap(records, lap_start, elapsed)

        cadence = first_number(lap, "avg_running_cadence", "avg_cadence")
        fractional_cadence = first_number(lap, "avg_fractional_cadence") or 0
        cadence_spm = (float(cadence) + float(fractional_cadence)) * 2 if cadence is not None else None

        avg_heart_rate = first_number(lap, "avg_heart_rate")
        if avg_heart_rate is None:
            avg_heart_rate = average([record.get("heart_rate") for record in lap_records], 1)
        max_heart_rate = first_number(lap, "max_heart_rate")
        if max_heart_rate is None:
            max_heart_rate = maximum([record.get("heart_rate") for record in lap_records])
        avg_power = first_number(lap, "avg_power")
        if avg_power is None:
            avg_power = average([record.get("power") for record in lap_records], 1)

        start_stamina = None
        end_stamina = None
        if stamina_supported:
            start_stamina = stamina_at(lap_records)
            end_stamina = stamina_at(reversed(lap_records))
            if index == 1:
                start_stamina = first_number(session, STAMINA_SESSION_START) or start_stamina
            if index == len(laps):
                end_stamina = first_number(session, *STAMINA_SESSION_END_FIELDS) or end_stamina

        result.append(
            {
                "split_index": index,
                "split_distance_m": round(distance, 3) if distance else None,
                "elapsed_time_sec": int_or_none(elapsed),
                "avg_hr": int_or_none(avg_heart_rate),
                "max_hr": int_or_none(max_heart_rate),
                "avg_power_w": int_or_none(avg_power),
                "avg_cadence_spm": float_or_none(rounded(cadence_spm, 1)),
                "avg_stride_length_mm": float_or_none(rounded(lap.get("avg_step_length"), 1)),
                "avg_gct_ms": float_or_none(rounded(lap.get("avg_stance_time"), 1)),
                "avg_vertical_ratio_pct": float_or_none(rounded(lap.get("avg_vertical_ratio"), 1)),
                "avg_vertical_oscillation_mm": float_or_none(rounded(lap.get("avg_vertical_oscillation"), 1)),
                "elevation_gain_m": float_or_none(rounded(lap.get("total_ascent"), 1)),
                "elevation_loss_m": float_or_none(rounded(lap.get("total_descent"), 1)),
                "stamina_start_pct": int_or_none(start_stamina),
                "stamina_end_pct": int_or_none(end_stamina),
            }
        )
    return result


def fit_text_value(value):
    if isinstance(value, (list, tuple)):
        for item in value:
            text = str(item or "").strip()
            if text:
                return text
        return None
    text = str(value or "").strip()
    return text or None


def sqlite_workout_structure_row(messages):
    workout_messages = messages.get("workout_mesgs", []) if messages else []
    workout_steps = messages.get("workout_step_mesgs", []) if messages else []
    workout_splits = messages.get("split_mesgs", []) if messages else []
    workout = workout_messages[0] if workout_messages else {}
    has_structure = bool(workout_messages or workout_steps or workout_splits)
    return {
        "has_workout_structure": 1 if has_structure else 0,
        "source": "fit",
        "sport": null_if_blank(workout.get("sport")),
        "sub_sport": null_if_blank(workout.get("sub_sport")),
        "workout_name": fit_text_value(workout.get("wkt_name")),
        "workout_description": fit_text_value(workout.get("wkt_description")),
        "num_valid_steps": int_or_none(workout.get("num_valid_steps")),
    }


def sqlite_workout_step_rows(messages):
    workout_steps = messages.get("workout_step_mesgs", []) if messages else []
    rows = []
    for index, step in enumerate(workout_steps, start=1):
        rows.append(
            {
                "step_index": index,
                "source_message_index": int_or_none(step.get("message_index")),
                "intensity": null_if_blank(step.get("intensity")),
                "duration_type": null_if_blank(step.get("duration_type")),
                "duration_value": int_or_none(step.get("duration_value")),
                "duration_distance_m": float_or_none(step.get("duration_distance")),
                "duration_time_sec": float_or_none(step.get("duration_time")),
                "target_type": null_if_blank(step.get("target_type")),
                "target_value": float_or_none(step.get("target_value")),
                "target_value_low": float_or_none(step.get("target_value_low")),
                "target_value_high": float_or_none(step.get("target_value_high")),
                "target_hr_zone": int_or_none(step.get("target_hr_zone")),
                "repeat_steps": int_or_none(step.get("repeat_steps")),
                "secondary_target_value": float_or_none(step.get("secondary_target_value")),
                "custom_target_value_low": float_or_none(step.get("custom_target_value_low")),
                "custom_target_value_high": float_or_none(step.get("custom_target_value_high")),
            }
        )
    return rows


def sqlite_workout_split_rows(messages):
    workout_splits = messages.get("split_mesgs", []) if messages else []
    rows = []
    for index, split in enumerate(workout_splits, start=1):
        rows.append(
            {
                "split_index": index,
                "source_message_index": int_or_none(split.get("message_index")),
                "split_type": null_if_blank(split.get("split_type")),
                "num_splits": int_or_none(split.get("num_splits")),
                "total_distance_m": float_or_none(split.get("total_distance")),
                "total_timer_time_sec": float_or_none(split.get("total_timer_time")),
                "avg_speed_mps": float_or_none(split.get("avg_speed")),
                "sport": null_if_blank(split.get("sport")),
                "sub_sport": null_if_blank(split.get("sub_sport")),
            }
        )
    return rows


def sync_activity_workout_structure(connection, activity_id, messages):
    structure_row = sqlite_workout_structure_row(messages)
    workout_step_rows = sqlite_workout_step_rows(messages)
    workout_split_rows = sqlite_workout_split_rows(messages)
    columns = [
        "activity_id",
        "has_workout_structure",
        "source",
        "sport",
        "sub_sport",
        "workout_name",
        "workout_description",
        "num_valid_steps",
    ]
    connection.execute(
        """
        INSERT INTO activity_workout_structure (
            activity_id,
            has_workout_structure,
            source,
            sport,
            sub_sport,
            workout_name,
            workout_description,
            num_valid_steps,
            updated_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
        ON CONFLICT(activity_id) DO UPDATE SET
            has_workout_structure = excluded.has_workout_structure,
            source = excluded.source,
            sport = excluded.sport,
            sub_sport = excluded.sub_sport,
            workout_name = excluded.workout_name,
            workout_description = excluded.workout_description,
            num_valid_steps = excluded.num_valid_steps,
            updated_at = CURRENT_TIMESTAMP
        """,
        [activity_id, *[structure_row[column] for column in columns[1:]]],
    )
    connection.execute("DELETE FROM activity_workout_step WHERE activity_id = ?", (activity_id,))
    connection.execute("DELETE FROM activity_workout_split WHERE activity_id = ?", (activity_id,))

    step_columns = [
        "activity_id",
        "step_index",
        "source_message_index",
        "intensity",
        "duration_type",
        "duration_value",
        "duration_distance_m",
        "duration_time_sec",
        "target_type",
        "target_value",
        "target_value_low",
        "target_value_high",
        "target_hr_zone",
        "repeat_steps",
        "secondary_target_value",
        "custom_target_value_low",
        "custom_target_value_high",
    ]
    split_columns = [
        "activity_id",
        "split_index",
        "source_message_index",
        "split_type",
        "num_splits",
        "total_distance_m",
        "total_timer_time_sec",
        "avg_speed_mps",
        "sport",
        "sub_sport",
    ]
    for step_row in workout_step_rows:
        row = {"activity_id": activity_id, **step_row}
        placeholders = ", ".join("?" for _ in step_columns)
        connection.execute(
            f"INSERT INTO activity_workout_step ({', '.join(step_columns)}) VALUES ({placeholders})",
            [row[column] for column in step_columns],
        )
    for split_row in workout_split_rows:
        row = {"activity_id": activity_id, **split_row}
        placeholders = ", ".join("?" for _ in split_columns)
        connection.execute(
            f"INSERT INTO activity_workout_split ({', '.join(split_columns)}) VALUES ({placeholders})",
            [row[column] for column in split_columns],
        )
    return {
        "structure": structure_row,
        "step_count": len(workout_step_rows),
        "split_count": len(workout_split_rows),
    }


def ensure_sqlite_schema(connection, schema_path=SQLITE_SCHEMA_PATH):
    connection.execute("PRAGMA foreign_keys = ON")
    ensure_activity_gps_columns(connection)
    for view_name in (
        "activity_view",
        "kilometer_split_view",
        "activity_training_purpose_view",
        "shoe_statistics_view",
    ):
        connection.execute(f"DROP VIEW IF EXISTS {view_name}")
    connection.executescript(schema_path.read_text(encoding="utf-8"))


def ensure_activity_gps_columns(connection):
    table_exists = connection.execute(
        "SELECT 1 FROM sqlite_master WHERE type = 'table' AND name = 'activity'"
    ).fetchone()
    if not table_exists:
        return
    existing_columns = {
        row[1]
        for row in connection.execute("PRAGMA table_info(activity)").fetchall()
    }
    for column_name in ("start_latitude", "start_longitude", "end_latitude", "end_longitude"):
        if column_name in existing_columns:
            continue
        connection.execute(f"ALTER TABLE activity ADD COLUMN {column_name} REAL")


def ensure_dropdown_source_table(connection, dropdown_options):
    connection.execute(
        f"""
        CREATE TABLE IF NOT EXISTS {DROPDOWN_SOURCE_TABLE} (
            source_key TEXT PRIMARY KEY,
            source_json TEXT NOT NULL,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    payloads = {
        "shoes": dropdown_options.get("shoes", []),
        "workout_types": dropdown_options.get("workout_types", []),
        "training_focus": dropdown_options.get("training_focus", []),
        "garmin_rpe": dropdown_options.get("garmin_rpe", []),
        "garmin_feel": dropdown_options.get("garmin_feel", []),
        "workout_focus_map": dropdown_options.get("workout_focus_map", {}),
    }
    for source_key, value in payloads.items():
        existing = connection.execute(
            f"SELECT source_json FROM {DROPDOWN_SOURCE_TABLE} WHERE source_key = ?",
            (source_key,),
        ).fetchone()
        if existing:
            continue
        connection.execute(
            f"""
            INSERT INTO {DROPDOWN_SOURCE_TABLE} (source_key, source_json, updated_at)
            VALUES (?, ?, CURRENT_TIMESTAMP)
            """,
            (source_key, json.dumps(value, ensure_ascii=False)),
        )


def ensure_feedback_dictionary_table(connection, dropdown_options):
    connection.execute(
        f"""
        CREATE TABLE IF NOT EXISTS {FEEDBACK_DICTIONARY_TABLE} (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            dictionary_key TEXT NOT NULL,
            label TEXT NOT NULL,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(dictionary_key, label)
        )
        """
    )
    seeds = load_feedback_dictionary_seed()
    for dictionary_key in ("garmin_rpe", "garmin_feel"):
        existing_count = connection.execute(
            f"SELECT COUNT(*) FROM {FEEDBACK_DICTIONARY_TABLE} WHERE dictionary_key = ?",
            (dictionary_key,),
        ).fetchone()[0]
        if existing_count:
            continue
        values = dropdown_options.get(dictionary_key) or seeds.get(dictionary_key, [])
        for label in values:
            value = str(label or "").strip()
            if not value:
                continue
            connection.execute(
                f"""
                INSERT INTO {FEEDBACK_DICTIONARY_TABLE} (dictionary_key, label, updated_at)
                VALUES (?, ?, CURRENT_TIMESTAMP)
                """,
                (dictionary_key, value),
            )


def ensure_workout_purpose_map_table(connection, dropdown_options):
    connection.execute(
        """
        CREATE TABLE IF NOT EXISTS workout_type_training_purpose_map (
            workout_type_id INTEGER PRIMARY KEY,
            primary_training_purpose_id INTEGER,
            secondary_training_purpose_id INTEGER,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (workout_type_id) REFERENCES workout_type(id),
            FOREIGN KEY (primary_training_purpose_id) REFERENCES training_purpose(id),
            FOREIGN KEY (secondary_training_purpose_id) REFERENCES training_purpose(id)
        )
        """
    )
    existing_count = connection.execute(
        "SELECT COUNT(*) FROM workout_type_training_purpose_map"
    ).fetchone()[0]
    if existing_count:
        return
    focus_map = dropdown_options.get("workout_focus_map", {})
    for workout_label in dropdown_options.get("workout_types", []):
        workout_row = workout_type_dimension_row(workout_label)
        workout_type_id = upsert_dimension(connection, "workout_type", "workout_type_code", workout_row)
        purpose_labels = focus_map.get(workout_label, [])
        if not purpose_labels:
            continue
        primary_id = None
        secondary_id = None
        if len(purpose_labels) >= 1:
            primary_id = upsert_dimension(
                connection,
                "training_purpose",
                "training_purpose_code",
                training_purpose_dimension_row(purpose_labels[0]),
            )
        if len(purpose_labels) >= 2:
            secondary_id = upsert_dimension(
                connection,
                "training_purpose",
                "training_purpose_code",
                training_purpose_dimension_row(purpose_labels[1]),
            )
        connection.execute(
            """
            INSERT INTO workout_type_training_purpose_map (
                workout_type_id,
                primary_training_purpose_id,
                secondary_training_purpose_id,
                updated_at
            ) VALUES (?, ?, ?, CURRENT_TIMESTAMP)
            """,
            (workout_type_id, primary_id, secondary_id),
        )


def upsert_dimension(connection, table, code_column, row):
    existing = connection.execute(
        f"SELECT id FROM {table} WHERE {code_column} = ?",
        (row[code_column],),
    ).fetchone()
    columns = list(row)
    if existing:
        update_columns = [column for column in columns if column != code_column]
        assignments = ", ".join(f"{column} = ?" for column in update_columns)
        values = [row[column] for column in update_columns]
        values.append(row[code_column])
        connection.execute(
            f"UPDATE {table} SET {assignments} WHERE {code_column} = ?",
            values,
        )
        return existing[0]

    placeholders = ", ".join("?" for _ in columns)
    connection.execute(
        f"INSERT INTO {table} ({', '.join(columns)}) VALUES ({placeholders})",
        [row[column] for column in columns],
    )
    return connection.execute(
        f"SELECT id FROM {table} WHERE {code_column} = ?",
        (row[code_column],),
    ).fetchone()[0]


def seed_reference_data(connection, dropdown_options):
    for label in dropdown_options.get("shoes", []):
        upsert_dimension(connection, "shoe", "shoe_code", shoe_dimension_row(label))
    for label in dropdown_options.get("workout_types", []):
        upsert_dimension(connection, "workout_type", "workout_type_code", workout_type_dimension_row(label))
    for label in dropdown_options.get("training_focus", []):
        upsert_dimension(connection, "training_purpose", "training_purpose_code", training_purpose_dimension_row(label))


def bootstrap_sqlite_state(db_path=SQLITE_DB_PATH, dropdown_options=None):
    options = dropdown_options or load_dropdown_options(DROPDOWN_CONFIG_PATH)
    db_path.parent.mkdir(parents=True, exist_ok=True)
    with sqlite3.connect(db_path) as connection:
        ensure_sqlite_schema(connection)
        seed_reference_data(connection, options)
        ensure_dropdown_source_table(connection, options)
        ensure_feedback_dictionary_table(connection, options)
        ensure_workout_purpose_map_table(connection, options)
    return db_path


def resolve_shoe_id(connection, metadata):
    label = null_if_blank(metadata.get("shoe"))
    if not label:
        return None
    return upsert_dimension(connection, "shoe", "shoe_code", shoe_dimension_row(label))


def resolve_workout_type_id(connection, metadata):
    label = null_if_blank(metadata.get("workout_type"))
    if not label:
        return None
    return upsert_dimension(connection, "workout_type", "workout_type_code", workout_type_dimension_row(label))


def metadata_training_focus_values(metadata, dropdown_options):
    focus = null_if_blank(metadata.get("training_focus"))
    if focus:
        values = []
        seen = set()
        for value in re.split(r"[、,;/\n]+", focus):
            label = str(value or "").strip()
            if not label or label in seen:
                continue
            seen.add(label)
            values.append(label)
        return values
    workout_type = null_if_blank(metadata.get("workout_type"))
    if not workout_type:
        return []
    focus_map = dropdown_options.get("workout_focus_map", {})
    return focus_map.get(workout_type, []) or focus_map.get(label_primary(workout_type), [])


def sync_activity_training_purposes(connection, activity_id, metadata, dropdown_options):
    connection.execute(
        "DELETE FROM activity_training_purpose WHERE activity_id = ?",
        (activity_id,),
    )
    seen_codes = set()
    for index, label in enumerate(metadata_training_focus_values(metadata, dropdown_options)):
        row = training_purpose_dimension_row(label)
        if row["training_purpose_code"] in seen_codes:
            continue
        seen_codes.add(row["training_purpose_code"])
        training_purpose_id = upsert_dimension(
            connection,
            "training_purpose",
            "training_purpose_code",
            row,
        )
        connection.execute(
            """
            INSERT INTO activity_training_purpose (
                activity_id,
                training_purpose_id,
                purpose_role
            ) VALUES (?, ?, ?)
            """,
            (activity_id, training_purpose_id, "PRIMARY" if index == 0 else "SECONDARY"),
        )


def upsert_activity(connection, row):
    columns = list(row)
    existing = connection.execute(
        "SELECT id FROM activity WHERE fit_sha256 = ?",
        (row["fit_sha256"],),
    ).fetchone()
    if existing:
        update_columns = [column for column in columns if column != "fit_sha256"]
        assignments = ", ".join(f"{column} = ?" for column in update_columns)
        values = [row[column] for column in update_columns]
        values.append(row["fit_sha256"])
        connection.execute(
            f"UPDATE activity SET {assignments} WHERE fit_sha256 = ?",
            values,
        )
        return existing[0]

    placeholders = ", ".join("?" for _ in columns)
    connection.execute(
        f"INSERT INTO activity ({', '.join(columns)}) VALUES ({placeholders})",
        [row[column] for column in columns],
    )
    return connection.execute(
        "SELECT id FROM activity WHERE fit_sha256 = ?",
        (row["fit_sha256"],),
    ).fetchone()[0]


def find_existing_activity_id(connection, fit_path: Path, activity_row):
    candidates = []
    fit_sha = activity_row.get("fit_sha256")
    garmin_activity_id = activity_row.get("garmin_activity_id")
    source_file_name = fit_path.name
    if fit_sha:
        candidates.append(
            connection.execute(
                "SELECT id FROM activity WHERE fit_sha256 = ?",
                (fit_sha,),
            ).fetchone()
        )
    if garmin_activity_id:
        candidates.append(
            connection.execute(
                "SELECT id FROM activity WHERE garmin_activity_id = ?",
                (garmin_activity_id,),
            ).fetchone()
        )
    candidates.append(
        connection.execute(
            "SELECT id FROM activity WHERE source_file_name = ?",
            (source_file_name,),
        ).fetchone()
    )
    for row in candidates:
        if row:
            return row[0]
    return None


def write_fit_to_sqlite(fit_path: Path, db_path: Path, metadata=None, fetch_weather=True, dropdown_options=None):
    dropdown_options = dropdown_options or load_dropdown_options()
    messages = decode_fit(fit_path)
    rows, session = build_rows(messages)
    if not rows:
        raise RuntimeError("No lap data found in FIT file.")
    metadata = finalized_metadata(metadata or {}, messages, session, rows, fit_path, fetch_weather, dropdown_options)

    activity_row = sqlite_activity_row(fit_path, messages, session, rows, metadata)
    split_rows = sqlite_split_rows(messages)

    bootstrap_sqlite_state(db_path, dropdown_options)
    with sqlite3.connect(db_path) as connection:
        activity_row["shoe_id"] = resolve_shoe_id(connection, metadata)
        activity_row["workout_type_id"] = resolve_workout_type_id(connection, metadata)
        activity_id = upsert_activity(connection, activity_row)
        sync_activity_training_purposes(connection, activity_id, metadata, dropdown_options)
        connection.execute("DELETE FROM kilometer_split WHERE activity_id = ?", (activity_id,))
        split_columns = [
            "activity_id",
            "split_index",
            "split_distance_m",
            "elapsed_time_sec",
            "avg_hr",
            "max_hr",
            "avg_power_w",
            "avg_cadence_spm",
            "avg_stride_length_mm",
            "avg_gct_ms",
            "avg_vertical_ratio_pct",
            "avg_vertical_oscillation_mm",
            "elevation_gain_m",
            "elevation_loss_m",
            "stamina_start_pct",
            "stamina_end_pct",
        ]
        placeholders = ", ".join("?" for _ in split_columns)
        for split_row in split_rows:
            row = {"activity_id": activity_id, **split_row}
            connection.execute(
                f"INSERT INTO kilometer_split ({', '.join(split_columns)}) VALUES ({placeholders})",
                [row[column] for column in split_columns],
            )
        sync_activity_workout_structure(connection, activity_id, messages)
    return db_path


def refresh_sqlite_gps_and_workout_structure(fit_path: Path, db_path: Path, metadata=None, fetch_weather=True, dropdown_options=None):
    dropdown_options = dropdown_options or load_dropdown_options()
    messages = decode_fit(fit_path)
    rows, session = build_rows(messages)
    if not rows:
        raise RuntimeError("No lap data found in FIT file.")
    metadata = finalized_metadata(metadata or {}, messages, session, rows, fit_path, fetch_weather, dropdown_options)
    activity_row = sqlite_activity_row(fit_path, messages, session, rows, metadata)

    bootstrap_sqlite_state(db_path, dropdown_options)
    with sqlite3.connect(db_path) as connection:
        activity_id = find_existing_activity_id(connection, fit_path, activity_row)
        if activity_id is None:
            return {
                "status": "missing_activity",
                "activity_id": None,
                "gps_updated": False,
                "workout_structure_updated": False,
            }
        connection.execute(
            """
            UPDATE activity
            SET fit_sha256 = ?,
                garmin_activity_id = COALESCE(garmin_activity_id, ?),
                source_file_name = ?,
                start_latitude = ?,
                start_longitude = ?,
                end_latitude = ?,
                end_longitude = ?,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (
                activity_row.get("fit_sha256"),
                activity_row.get("garmin_activity_id"),
                fit_path.name,
                activity_row.get("start_latitude"),
                activity_row.get("start_longitude"),
                activity_row.get("end_latitude"),
                activity_row.get("end_longitude"),
                activity_id,
            ),
        )
        sync_activity_workout_structure(connection, activity_id, messages)
    return {
        "status": "updated",
        "activity_id": activity_id,
        "gps_updated": True,
        "workout_structure_updated": True,
    }


def create_workbook(fit_path: Path, output_path: Path, metadata=None, fetch_weather=True, dropdown_options=None):
    dropdown_options = dropdown_options or load_dropdown_options()
    messages = decode_fit(fit_path)
    rows, session = build_rows(messages)
    if not rows:
        raise RuntimeError("No lap data found in FIT file.")
    metadata = finalized_metadata(metadata or {}, messages, session, rows, fit_path, fetch_weather, dropdown_options)

    wb = Workbook()
    ws = wb.active
    ws.title = "每公里數據"
    add_metadata_sheet(wb, metadata, fit_path, session, rows, messages.get("record_mesgs", []), dropdown_options)
    add_workout_structure_sheet(wb, messages)
    date_label = activity_date(session, fit_path)
    ws["A1"] = f"{WORKBOOK_VERSION_NAME} - {date_label} (資料來源: {fit_path.name})"
    ws.append(HEADERS)
    for row in rows:
        ws.append(row)
    add_percentage_values(ws, len(rows), metadata)
    total_row = add_total_row(ws, len(rows), rows)
    apply_styles(ws, total_row)
    add_charts(wb, len(rows))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def main():
    parser = argparse.ArgumentParser(description="Convert a Garmin FIT running activity to a per-kilometer Excel workbook.")
    parser.add_argument("--version", action="version", version=f"Running Analytics v{APP_VERSION} / {WORKBOOK_VERSION_NAME}")
    parser.add_argument("fit_file", type=Path)
    parser.add_argument("-o", "--output", type=Path)
    parser.add_argument("--sqlite-db", type=Path, help="Also import this FIT activity into a SQLite v1.0 database.")
    parser.add_argument("--interactive", action="store_true", help="Prompt for manual activity metadata before exporting.")
    parser.set_defaults(fetch_weather=True)
    parser.add_argument("--fetch-weather", dest="fetch_weather", action="store_true", help="Fetch weather from Open-Meteo using FIT start time and GPS location. Enabled by default.")
    parser.add_argument("--no-fetch-weather", dest="fetch_weather", action="store_false", help="Skip automatic weather lookup.")
    parser.add_argument("--dropdown-config", type=Path, default=DROPDOWN_CONFIG_PATH, help="JSON file for dropdown options.")
    parser.add_argument("--activity-name", help="Activity name shown in the Activity section.")
    parser.add_argument("--shoe", help="Shoe name, e.g. 'Adidas Boston 13'.")
    parser.add_argument("--weather-temp", type=float, help="Weather temperature in Celsius.")
    parser.add_argument("--humidity", type=float, help="Humidity percentage.")
    parser.add_argument("--wind-direction", help="Wind direction.")
    parser.add_argument("--wind-speed", help="Wind speed, e.g. '12 km/h'.")
    parser.add_argument("--weather-description", help="Weather description, e.g. sunny, cloudy, light rain.")
    parser.add_argument("--workout-type", help="Workout type, e.g. Recovery, Tempo, LSD, Intervals.")
    parser.add_argument("--training-focus", help="Training focus, e.g. Aerobic, Threshold, VO2max.")
    parser.add_argument("--feel", help="Workout feel, e.g. 50 or '普通'.")
    parser.add_argument("--rpe", help="Effort rating, e.g. 3, 30, or '3 - 中等'.")
    parser.add_argument("--fueling", help="Free-form fueling notes.")
    parser.add_argument("--max-hr", type=float, help="Maximum heart rate used for average heart rate percentage.")
    parser.add_argument("--critical-power", type=float, help="Critical Power in watts used for average power percentage.")
    parser.add_argument("--training-effect-aerobic", type=float, help="Aerobic Training Effect summary.")
    parser.add_argument("--training-effect-anaerobic", type=float, help="Anaerobic Training Effect summary.")
    parser.add_argument("--training-load", type=float, help="Training Load summary.")
    parser.add_argument("--recovery-time-hr", type=float, help="Recovery time in hours.")
    parser.add_argument("--notes", help="Free-form notes.")
    args = parser.parse_args()

    output = args.output
    if output is None:
        output = default_output_path(args.fit_file)

    dropdown_options = load_dropdown_options(args.dropdown_config)
    metadata = collect_metadata(args, dropdown_options)
    saved = create_workbook(
        args.fit_file,
        output,
        metadata,
        fetch_weather=args.fetch_weather,
        dropdown_options=dropdown_options,
    )
    print(saved)
    if args.sqlite_db:
        db_path = write_fit_to_sqlite(
            args.fit_file,
            args.sqlite_db,
            metadata,
            fetch_weather=args.fetch_weather,
            dropdown_options=dropdown_options,
        )
        print(db_path)


if __name__ == "__main__":
    main()
